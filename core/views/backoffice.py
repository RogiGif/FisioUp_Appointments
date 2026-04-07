from datetime import datetime, timedelta, time as dtime
from decimal import Decimal
from collections import defaultdict
from dataclasses import dataclass
import base64
import binascii
import json
import csv
import io
import unicodedata

from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User, Group
from django.contrib.auth.forms import PasswordChangeForm, SetPasswordForm
from django.contrib.auth import views as auth_views
from django.core.cache import cache
from django.core.files.base import ContentFile
from django.core.exceptions import PermissionDenied
from django.core.paginator import Paginator
from django.db import IntegrityError, transaction
from django.db.models import Q, Count, Sum, Exists, OuterRef, F
from django.db.models.functions import Coalesce, TruncDate
from django.shortcuts import render, redirect, get_object_or_404
from django.conf import settings
from django.urls import reverse
from urllib.parse import urlencode
from django.utils.text import slugify
from django.utils import timezone
from django.utils.functional import cached_property
from django.utils.http import url_has_allowed_host_and_scheme
from django.utils.safestring import mark_safe
from django.http import HttpResponse, HttpResponseForbidden, JsonResponse
from django.views.decorators.http import require_POST, require_http_methods, require_GET

from core.decorators import professional_required, backoffice_required
from core.permissions import (
    can_view_all_calendar,
    can_book_for_any_professional,
    can_access_backoffice,
    is_admin_role,
    is_receptionist,
)
from core.ratelimit import check_rate_limit, rate_limited_response, is_json_request, rate_limit
from core.emails import send_templated_email, clinic_email, clinic_settings, log_email_skip
from core.forms import (
    RegisterForm,
    ClientProfileForm,
    ProfessionalProfileForm,
    StaffClientCreateForm,
    BackofficeServiceForm,
    BackofficeProfessionalForm,
    BackofficePartnerForm,
    BackofficeHighlightForm,
    ClinicEmailSettingsForm,
    WeeklyScheduleForm,
    WeeklyWorkingBlockFormSet,
    WeeklyBreakBlockFormSet,
    CashSessionOpenForm,
    CashSessionCloseForm,
    CashManualMovementForm,
    CashAppointmentMovementForm,
    CashClientPaymentMovementForm,
    CashGroupMonthlyMovementForm,
    CashVoidMovementForm,
    MoloniCustomerDefaultsForm,
)
from core.utils.pricing import (
    compute_pricing,
    recalculate_upcoming_appointment_prices,
    recalculate_partner_upcoming_appointments,
)
from core.services.subcontracting import sync_subcontractor_payout
from core.services.payments import ensure_client_payment_cash_movement
from core.services.audit import log_audit_event, snapshot_instance, cleanup_old_audit_logs_if_needed
from core.services import moloni as moloni_service
from core.services.moloni_sync import (
    apply_remote_customer_to_profile as moloni_apply_remote_customer_to_profile,
    build_reconciliation_report as moloni_build_reconciliation_report,
    run_bidirectional_reconciliation as moloni_run_bidirectional_reconciliation,
    sync_customers as moloni_sync_customers,
)
from core.utils.revenue import (
    get_revenue_queryset,
    compute_trend,
    month_range,
    week_range,
    day_range,
    month_start,
)
from core.views.common import _find_matching_cancelled_appointment, log_appt
from core.models import (
    Professional,
    WeeklySchedule,
    WeeklyWorkingBlock,
    WeeklyBreakBlock,
    Appointment,
    Service,
    ClientProfile,
    ClinicalRecord,
    TreatmentRecord,
    AppointmentLog,
    BlockedSlot,
    GroupSession,
    GroupEnrollment,
    GroupMonthlyCharge,
    MoloniIntegration,
    ClientImportLog,
    ClientImportBatch,
    ClientImportRow,
    Partner,
    PartnerServicePrice,
    Product,
    StockMovement,
    ContentPost,
    SubcontractorPaymentLine,
    ClinicSettings,
    AuditLog,
    CashSession,
    CashMovement,
    ClientPayment,
)

from core.views.common import *


def test_duralux(request):
    return render(request, "core/base_duralux.html")


def _user_can_use_post_designer(user):
    if not user.is_authenticated:
        return False
    return Professional.objects.filter(user=user).exists() or is_admin_role(user)


WEEKLY_SCHEDULE_AUDIT_FIELDS = ["id", "professional_id", "timezone", "is_active"]
WEEKLY_WORK_BLOCK_AUDIT_FIELDS = ["id", "weekly_schedule_id", "weekday", "start_time", "end_time", "location"]
WEEKLY_BREAK_BLOCK_AUDIT_FIELDS = ["id", "weekly_schedule_id", "weekday", "start_time", "end_time"]
PARTNER_AUDIT_FIELDS = [
    "id",
    "name",
    "active",
    "notes",
    "discount_type",
    "discount_percent",
    "discount_amount",
    "discount_label",
    "logo",
]
PARTNER_SERVICE_PRICE_AUDIT_FIELDS = [
    "id",
    "partner_id",
    "service_id",
    "price",
    "pricing_mode",
    "price_first",
    "price_followup",
    "discount_type",
    "discount_percent",
    "discount_amount",
    "is_enabled",
]
CLIENT_IMPORT_BATCH_AUDIT_FIELDS = ["id", "uploaded_by_id", "original_filename", "validate_nif", "created_at"]
CLIENT_IMPORT_LOG_AUDIT_FIELDS = [
    "id",
    "created_by_id",
    "file_name",
    "created_count",
    "updated_count",
    "skipped_count",
    "error_count",
    "summary",
    "created_at",
]


def _weekly_schedule_snapshot(schedule):
    return {
        "schedule": snapshot_instance(schedule, WEEKLY_SCHEDULE_AUDIT_FIELDS),
        "blocks": [
            snapshot_instance(block, WEEKLY_WORK_BLOCK_AUDIT_FIELDS)
            for block in schedule.blocks.order_by("weekday", "start_time", "id")
        ],
        "breaks": [
            snapshot_instance(block, WEEKLY_BREAK_BLOCK_AUDIT_FIELDS)
            for block in schedule.breaks.order_by("weekday", "start_time", "id")
        ],
    }


def _partner_snapshot(partner):
    return snapshot_instance(partner, PARTNER_AUDIT_FIELDS)


def _partner_prices_snapshot(partner):
    return [
        snapshot_instance(item, PARTNER_SERVICE_PRICE_AUDIT_FIELDS)
        for item in PartnerServicePrice.objects.filter(partner=partner).select_related("service").order_by("service__name", "id")
    ]


def _client_import_batch_snapshot(batch):
    return snapshot_instance(batch, CLIENT_IMPORT_BATCH_AUDIT_FIELDS)


def _client_import_log_snapshot(log):
    return snapshot_instance(log, CLIENT_IMPORT_LOG_AUDIT_FIELDS)


def _audit_log_object_url(log):
    if not log.content_type_id or not log.object_id:
        return ""

    model = log.content_type.model
    try:
        if model == "appointment":
            return reverse("professional_appointment_detail", args=[log.object_id])
        if model == "clientprofile":
            return reverse("prof_customer_detail", kwargs={"client_id": log.object_id})
        if model == "professional":
            return reverse("backoffice_professional_edit", kwargs={"professional_id": log.object_id})
        if model == "weeklyschedule":
            professional_id = (
                ((log.after or {}).get("schedule") or {}).get("professional_id")
                or ((log.before or {}).get("schedule") or {}).get("professional_id")
            )
            if professional_id:
                return reverse("backoffice_weekly_schedule_edit", kwargs={"professional_id": professional_id})
        if model == "partner":
            return reverse("backoffice_partner_edit", kwargs={"partner_id": log.object_id})
        if model == "contentpost":
            return reverse("backoffice_highlight_edit", kwargs={"post_id": log.object_id})
        if model == "groupschedule":
            return reverse("group_schedule_edit", kwargs={"schedule_id": log.object_id})
        if model == "groupsession":
            return reverse("group_session_detail_admin", kwargs={"session_id": log.object_id})
        if model == "groupenrollment":
            session_id = (
                (log.after or {}).get("session_id")
                or (log.before or {}).get("session_id")
            )
            if session_id:
                return reverse("group_session_detail_admin", kwargs={"session_id": session_id})
        if model == "product":
            return reverse("backoffice_stock_product_edit", kwargs={"product_id": log.object_id})
        if model == "productcategory":
            return reverse("backoffice_stock_category_edit", kwargs={"category_id": log.object_id})
        if model == "cashsession":
            return f"{reverse('backoffice_cash_dashboard')}?session_id={log.object_id}"
        if model == "cashmovement":
            session_id = (log.after or {}).get("session_id") or (log.before or {}).get("session_id")
            if session_id:
                return f"{reverse('backoffice_cash_dashboard')}?session_id={session_id}"
            return reverse("backoffice_cash_dashboard")
        if model in {"clientimportbatch", "clientimportlog"}:
            return reverse("client_import")
        if model == "clinicalrecord":
            client_id = (log.after or {}).get("client_id") or (log.before or {}).get("client_id")
            if client_id:
                return reverse("prof_customer_detail", kwargs={"client_id": client_id})
        if model == "treatmentrecord":
            client_id = (log.after or {}).get("client_id") or (log.before or {}).get("client_id")
            if client_id:
                return reverse("prof_customer_detail", kwargs={"client_id": client_id})
        if model == "subcontractorpaymentline":
            return reverse("backoffice_subcontractors")
    except Exception:
        return ""


def _audit_log_pretty_payload(value):
    if value in (None, "", [], {}):
        return "-"
    try:
        return json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True)
    except TypeError:
        return str(value)


def _get_cached_audit_dashboard_data():
    cache_key = "audit_logs:dashboard:v2"
    cached = cache.get(cache_key)
    if cached:
        return cached

    now = timezone.now()
    data = {
        "summary_total": AuditLog.objects.count(),
        "summary_last_24h": AuditLog.objects.filter(created_at__gte=now - timedelta(hours=24)).count(),
        "summary_last_7d": AuditLog.objects.filter(created_at__gte=now - timedelta(days=7)).count(),
        "top_categories": list(
            AuditLog.objects.values("category")
            .annotate(total=Count("id"))
            .order_by("-total", "category")[:5]
        ),
        "top_actors": list(
            AuditLog.objects.exclude(actor_display="")
            .values("actor_display", "actor_role")
            .annotate(total=Count("id"))
            .order_by("-total", "actor_display")[:5]
        ),
        "object_type_choices": [
            f"{item['content_type__app_label']}.{item['content_type__model']}"
            for item in (
                AuditLog.objects
                .exclude(content_type__isnull=True)
                .values("content_type__app_label", "content_type__model")
                .order_by("content_type__app_label", "content_type__model")
                .distinct()
            )
        ],
        "category_choices": list(AuditLog.objects.order_by("category").values_list("category", flat=True).distinct()),
        "action_choices": list(AuditLog.objects.order_by("action").values_list("action", flat=True).distinct()),
        "source_choices": list(AuditLog.objects.exclude(source="").order_by("source").values_list("source", flat=True).distinct()),
        "role_choices": list(AuditLog.objects.exclude(actor_role="").order_by("actor_role").values_list("actor_role", flat=True).distinct()),
    }
    cache.set(cache_key, data, 300)
    return data


def _can_access_cash_area(user):
    return is_admin_role(user) or is_receptionist(user)


def _cash_summary_movements_queryset(qs):
    return qs.exclude(source_type=CashMovement.SOURCE_STOCK_SALE)


def _cash_session_totals(session):
    movements = _cash_summary_movements_queryset(session.movements.filter(is_void=False))
    all_movements = session.movements.all()
    total_in = (
        movements.filter(movement_type=CashMovement.TYPE_IN)
        .aggregate(total=Coalesce(Sum("amount"), Decimal("0.00")))
        .get("total")
        or Decimal("0.00")
    )
    total_out = (
        movements.filter(movement_type=CashMovement.TYPE_OUT)
        .aggregate(total=Coalesce(Sum("amount"), Decimal("0.00")))
        .get("total")
        or Decimal("0.00")
    )
    cash_in = (
        movements.filter(movement_type=CashMovement.TYPE_IN, payment_method=CashMovement.METHOD_CASH)
        .aggregate(total=Coalesce(Sum("amount"), Decimal("0.00")))
        .get("total")
        or Decimal("0.00")
    )
    cash_out = (
        movements.filter(movement_type=CashMovement.TYPE_OUT, payment_method=CashMovement.METHOD_CASH)
        .aggregate(total=Coalesce(Sum("amount"), Decimal("0.00")))
        .get("total")
        or Decimal("0.00")
    )
    expected_cash = (session.opening_amount or Decimal("0.00")) + cash_in - cash_out
    balance = (session.opening_amount or Decimal("0.00")) + total_in - total_out
    payment_breakdown = list(
        movements.values("payment_method")
        .annotate(
            total_in=Coalesce(Sum("amount", filter=Q(movement_type=CashMovement.TYPE_IN)), Decimal("0.00")),
            total_out=Coalesce(Sum("amount", filter=Q(movement_type=CashMovement.TYPE_OUT)), Decimal("0.00")),
        )
        .order_by("payment_method")
    )
    for row in payment_breakdown:
        row["net"] = (row["total_in"] or Decimal("0.00")) - (row["total_out"] or Decimal("0.00"))
        row["label"] = dict(CashMovement.PAYMENT_METHOD_CHOICES).get(row["payment_method"], row["payment_method"])
    source_breakdown = _cash_origin_breakdown(session)
    return {
        "total_in": total_in,
        "total_out": total_out,
        "balance": balance,
        "cash_in": cash_in,
        "cash_out": cash_out,
        "expected_cash": expected_cash,
        "movement_count": movements.count(),
        "voided_count": all_movements.filter(is_void=True).count(),
        "payment_breakdown": payment_breakdown,
        "source_breakdown": source_breakdown,
    }


def _cash_dashboard_querystring(request, selected_session, session_status, session_date_value, extra=None, remove=None):
    params = request.GET.copy()
    remove = set(remove or [])
    for key in remove:
        params.pop(key, None)
    if selected_session:
        params["session_id"] = str(selected_session.id)
    elif "session_id" in params:
        params.pop("session_id", None)
    if session_status and session_status != "all":
        params["session_status"] = session_status
    elif "session_status" in params:
        params.pop("session_status", None)
    if session_date_value:
        params["session_date"] = session_date_value
    elif "session_date" in params:
        params.pop("session_date", None)
    for key, value in (extra or {}).items():
        if value in (None, ""):
            params.pop(key, None)
        else:
            params[key] = str(value)
    return params.urlencode()


def _cash_origin_breakdown(session):
    movements = (
        _cash_summary_movements_queryset(session.movements.filter(is_void=False))
        .values("source_type")
        .annotate(total=Coalesce(Sum("amount"), Decimal("0.00")), count=Count("id"))
        .order_by("source_type")
    )
    labels = dict(CashMovement.SOURCE_CHOICES)
    rows = []
    for row in movements:
        rows.append({
            "source_type": row["source_type"],
            "label": labels.get(row["source_type"], row["source_type"]),
            "count": row["count"],
            "total": row["total"] or Decimal("0.00"),
        })
    return rows


def _cash_month_bounds(reference_date):
    month_start = reference_date.replace(day=1)
    next_month = (month_start + timedelta(days=32)).replace(day=1)
    month_end = next_month - timedelta(days=1)
    return month_start, month_end


def _appointment_charge_amount_for_report(appointment):
    if appointment.settlement_locked_at:
        return appointment.settlement_final_price or Decimal("0.00")
    return appointment.final_price or Decimal("0.00")


def _cash_month_summary(reference_date):
    month_start, month_end = _cash_month_bounds(reference_date)
    movement_qs = _cash_summary_movements_queryset(
        CashMovement.objects.filter(
            session__session_date__gte=month_start,
            session__session_date__lte=month_end,
            is_void=False,
        )
    )
    received_total = (
        movement_qs.filter(movement_type=CashMovement.TYPE_IN)
        .aggregate(total=Coalesce(Sum("amount"), Decimal("0.00")))
        .get("total")
        or Decimal("0.00")
    )
    expense_total = (
        movement_qs.filter(movement_type=CashMovement.TYPE_OUT)
        .aggregate(total=Coalesce(Sum("amount"), Decimal("0.00")))
        .get("total")
        or Decimal("0.00")
    )

    appointment_debts = []
    appointment_qs = (
        Appointment.objects
        .select_related("client", "client__client_profile", "service")
        .annotate(
            allocated_total=Coalesce(
                Sum(
                    "payment_allocations__allocated_amount",
                    filter=Q(payment_allocations__payment__status=ClientPayment.STATUS_POSTED),
                ),
                Decimal("0.00"),
            )
        )
        .filter(date__gte=month_start, date__lte=month_end)
        .exclude(status=Appointment.STATUS_CANCELLED)
        .order_by("date", "time", "id")
    )
    for appointment in appointment_qs:
        charge_amount = _appointment_charge_amount_for_report(appointment)
        paid_amount = appointment.allocated_total or Decimal("0.00")
        if not paid_amount and appointment.is_paid:
            paid_amount = charge_amount
        outstanding_amount = charge_amount - paid_amount
        if outstanding_amount <= 0:
            continue
        client_label = (
            getattr(getattr(appointment.client, "client_profile", None), "full_name", "")
            or appointment.client.get_full_name()
            or appointment.client.username
        )
        appointment_debts.append({
            "kind": "Marcação",
            "date": appointment.date,
            "time": appointment.time,
            "client_label": client_label,
            "description": getattr(appointment.service, "name", "") or "Serviço",
            "amount": outstanding_amount,
        })

    monthly_charge_debts = []
    monthly_charge_qs = (
        GroupMonthlyCharge.objects
        .select_related("client", "client__client_profile", "service")
        .annotate(
            allocated_total=Coalesce(
                Sum(
                    "payment_allocations__allocated_amount",
                    filter=Q(payment_allocations__payment__status=ClientPayment.STATUS_POSTED),
                ),
                Decimal("0.00"),
            )
        )
        .filter(month__gte=month_start, month__lte=month_end)
        .exclude(status=GroupMonthlyCharge.STATUS_VOID)
        .order_by("month", "id")
    )
    for charge in monthly_charge_qs:
        charge_amount = charge.final_price or Decimal("0.00")
        paid_amount = charge.allocated_total or Decimal("0.00")
        if not paid_amount and charge.status == GroupMonthlyCharge.STATUS_PAID:
            paid_amount = charge_amount
        outstanding_amount = charge_amount - paid_amount
        if outstanding_amount <= 0:
            continue
        client_label = (
            getattr(getattr(charge.client, "client_profile", None), "full_name", "")
            or charge.client.get_full_name()
            or charge.client.username
        )
        monthly_charge_debts.append({
            "kind": "Turma",
            "date": charge.month,
            "time": None,
            "client_label": client_label,
            "description": charge.class_name or getattr(charge.service, "name", "") or "Turma",
            "amount": outstanding_amount,
        })

    debt_rows = sorted(
        appointment_debts + monthly_charge_debts,
        key=lambda item: (item["date"], item["time"] or dtime.min, item["client_label"]),
    )
    debt_total = sum((row["amount"] for row in debt_rows), Decimal("0.00"))

    subcontract_qs = (
        SubcontractorPaymentLine.objects
        .select_related("professional", "professional__user", "client", "service")
        .filter(appointment_date__gte=month_start, appointment_date__lte=month_end)
        .exclude(status=SubcontractorPaymentLine.STATUS_VOID)
        .order_by("appointment_date", "appointment_time", "id")
    )
    subcontract_paid_total = (
        subcontract_qs.filter(status=SubcontractorPaymentLine.STATUS_PAID)
        .aggregate(total=Coalesce(Sum("payable_amount"), Decimal("0.00")))
        .get("total")
        or Decimal("0.00")
    )
    subcontract_open_total = (
        subcontract_qs.filter(status=SubcontractorPaymentLine.STATUS_UNPAID)
        .aggregate(total=Coalesce(Sum("payable_amount"), Decimal("0.00")))
        .get("total")
        or Decimal("0.00")
    )
    subcontract_preview = []
    for line in subcontract_qs.filter(status=SubcontractorPaymentLine.STATUS_UNPAID)[:8]:
        professional_label = line.professional.display_name if hasattr(line.professional, "display_name") else str(line.professional)
        subcontract_preview.append({
            "date": line.appointment_date,
            "time": line.appointment_time,
            "professional_label": professional_label,
            "client_label": line.client.full_name if line.client else "Utente",
            "description": getattr(line.service, "name", "") or "Serviço",
            "amount": line.payable_amount or Decimal("0.00"),
        })

    return {
        "month_start": month_start,
        "month_end": month_end,
        "session_count": CashSession.objects.filter(session_date__gte=month_start, session_date__lte=month_end).count(),
        "received_total": received_total,
        "expense_total": expense_total,
        "balance": received_total - expense_total,
        "debt_total": debt_total,
        "debt_count": len(debt_rows),
        "debt_preview": debt_rows[:8],
        "subcontract_paid_total": subcontract_paid_total,
        "subcontract_open_total": subcontract_open_total,
        "subcontract_open_count": subcontract_qs.filter(status=SubcontractorPaymentLine.STATUS_UNPAID).count(),
        "subcontract_preview": subcontract_preview,
    }


def _pending_cash_appointments_for_session(session):
    return (
        Appointment.objects.select_related("client", "client__client_profile", "service", "professional", "professional__user")
        .filter(is_paid=True)
        .filter(
            Q(paid_at__date=session.session_date)
            | Q(paid_at__isnull=True, date=session.session_date)
        )
        .exclude(status=Appointment.STATUS_CANCELLED)
        .exclude(final_price__lte=Decimal("0.00"))
        .filter(payment_allocations__isnull=True)
        .filter(cash_movement__isnull=True)
        .distinct()
        .order_by("time", "id")
    )


def _pending_cash_client_payments_for_session(session):
    return (
        ClientPayment.objects
        .select_related("client_profile", "created_by", "cash_movement")
        .filter(
            status=ClientPayment.STATUS_POSTED,
            received_at__date=session.session_date,
            cash_movement__isnull=True,
            amount_received__gt=Decimal("0.00"),
        )
        .order_by("received_at", "id")
    )


def _pending_cash_group_monthly_for_session(session):
    return (
        GroupMonthlyCharge.objects.select_related("client", "client__client_profile", "service", "professional", "professional__user", "schedule")
        .filter(status=GroupMonthlyCharge.STATUS_PAID)
        .filter(
            Q(paid_at__date=session.session_date)
            | Q(paid_at__isnull=True, month=session.session_date.replace(day=1))
        )
        .exclude(final_price__lte=Decimal("0.00"))
        .filter(payment_allocations__isnull=True)
        .filter(cash_movement__isnull=True)
        .distinct()
        .order_by("month", "id")
    )


@login_required(login_url="/login/")
def backoffice_post_designer_view(request):
    if not _user_can_use_post_designer(request.user):
        raise PermissionDenied("Acesso reservado a profissionais e administradores.")

    formats = [
        {"id": "landscape", "label": "Post horizontal", "width": 1200, "height": 846},
        {"id": "square", "label": "Post quadrado", "width": 1080, "height": 1080},
        {"id": "story", "label": "Story", "width": 1080, "height": 1920},
    ]
    presets = [
        {
            "id": "hero-left",
            "label": "Esquerda ampla",
            "title_box": {"x": 0.058, "y": 0.555, "w": 0.56, "align": "left"},
            "subtitle_box": {"x": 0.058, "y": 0.642, "w": 0.58, "align": "left"},
            "logo": {"x": 0.058, "y": 0.083, "width": 0.185},
        },
        {
            "id": "hero-bottom",
            "label": "Faixa inferior",
            "title_box": {"x": 0.067, "y": 0.733, "w": 0.56, "align": "left"},
            "subtitle_box": {"x": 0.067, "y": 0.815, "w": 0.58, "align": "left"},
            "logo": {"x": 0.067, "y": 0.083, "width": 0.185},
        },
        {
            "id": "hero-right",
            "label": "Direita",
            "title_box": {"x": 0.942, "y": 0.544, "w": 0.52, "align": "right"},
            "subtitle_box": {"x": 0.942, "y": 0.632, "w": 0.54, "align": "right"},
            "logo": {"x": 0.75, "y": 0.083, "width": 0.185},
        },
    ]

    ctx = {
        "is_post_designer_page": True,
        "post_designer_formats": formats,
        "post_designer_presets": presets,
        "post_designer_post_kinds": ContentPost.KIND_CHOICES,
        "post_designer_config": {
            "logoDefaultUrl": settings.STATIC_URL + "core/images/logo_fifioUP.svg",
            "logoWhiteUrl": settings.STATIC_URL + "core/images/logo_fifioUP_white.svg",
            "createHighlightUrl": reverse("backoffice_post_designer_create_highlight"),
            "formats": formats,
            "presets": presets,
        },
    }
    return render(request, "backoffice/post_designer.html", ctx)


@require_POST
@login_required(login_url="/login/")
def backoffice_post_designer_create_highlight_view(request):
    if not _user_can_use_post_designer(request.user):
        return JsonResponse({"ok": False, "error": "Sem permissão para criar destaques."}, status=403)

    title = (request.POST.get("title") or "").strip()
    subtitle = (request.POST.get("subtitle") or "").strip()
    body_text = (request.POST.get("body") or "").strip()
    kind = (request.POST.get("kind") or "news").strip().lower()
    image_data = (request.POST.get("image_data") or "").strip()
    image_format = (request.POST.get("format") or "landscape").strip().lower()

    if len(title) < 5:
        return JsonResponse({"ok": False, "error": "O título deve ter pelo menos 5 caracteres."}, status=400)

    if not image_data.startswith("data:image/"):
        return JsonResponse({"ok": False, "error": "Imagem final inválida."}, status=400)

    try:
        header, encoded = image_data.split(",", 1)
    except ValueError:
        return JsonResponse({"ok": False, "error": "Formato de imagem inválido."}, status=400)

    image_ext = "png"
    if ";base64" not in header:
        return JsonResponse({"ok": False, "error": "A imagem tem de ser enviada em base64."}, status=400)
    if "image/jpeg" in header:
        image_ext = "jpg"
    elif "image/webp" in header:
        image_ext = "webp"

    try:
        image_binary = base64.b64decode(encoded)
    except (binascii.Error, ValueError):
        return JsonResponse({"ok": False, "error": "Não foi possível ler a imagem exportada."}, status=400)

    valid_kinds = {choice[0] for choice in ContentPost.KIND_CHOICES}
    if kind not in valid_kinds:
        kind = "news"

    excerpt = subtitle or body_text[:180] or title
    body = body_text or subtitle or excerpt
    timestamp = timezone.now().strftime("%Y%m%d%H%M%S")
    filename = f"{slugify(title) or 'post'}-{image_format}-{timestamp}.{image_ext}"

    post = ContentPost(
        title=title,
        kind=kind,
        excerpt=excerpt,
        body=body,
        status="draft",
        is_featured=False,
        author=request.user,
    )
    post.cover_image.save(filename, ContentFile(image_binary), save=False)
    post.save()
    log_audit_event(
        category="content",
        action="draft_created",
        request=request,
        actor=request.user,
        instance=post,
        source="post_designer",
        message="Rascunho de destaque criado a partir do criador de posts.",
        after=snapshot_instance(
            post,
            fields=["title", "kind", "status", "slug", "author_id", "is_featured"],
        ),
        metadata={"format": image_format, "has_subtitle": bool(subtitle)},
    )

    payload = {
        "ok": True,
        "message": "Rascunho criado com sucesso.",
        "post_id": post.id,
        "redirect_url": reverse("backoffice_highlight_edit", args=[post.id]) if is_admin_role(request.user) else "",
    }
    if not is_admin_role(request.user):
        payload["message"] = "Rascunho criado com sucesso. Um administrador pode agora revê-lo em Destaques."
    return JsonResponse(payload)


def backoffice_dashboard_view(request):
    today = timezone.localdate()
    can_view_financial = is_admin_role(request.user)
    revenue_qs = get_revenue_queryset(request.user)
    can_view_all = can_view_all_calendar(request.user)
    prof = None
    if not can_view_all:
        prof = Professional.objects.filter(user=request.user).first()

    appt_metrics_qs = Appointment.objects.all()
    agenda_base_qs = Appointment.objects.select_related("client", "service", "professional", "professional__user")
    if prof:
        appt_metrics_qs = appt_metrics_qs.filter(professional=prof)
        agenda_base_qs = agenda_base_qs.filter(professional=prof)

    period = (request.GET.get("period") or "month").strip().lower()
    if period not in {"day", "week", "month"}:
        period = "month"

    today_start, today_end = day_range(today)
    yesterday_start, yesterday_end = day_range(today - timedelta(days=1))
    week_start_date, week_end_date = week_range(today)
    prev_week_start = week_start_date - timedelta(days=7)
    prev_week_end = week_start_date
    month_start_date, month_end_date = month_range(today)
    prev_month_end = month_start_date
    prev_month_start = month_start(prev_month_end - timedelta(days=1))

    if period == "day":
        period_start, period_end = today_start, today_end
        prev_period_start, prev_period_end = yesterday_start, yesterday_end
    elif period == "week":
        period_start, period_end = week_start_date, week_end_date
        prev_period_start, prev_period_end = prev_week_start, prev_week_end
    else:
        period_start, period_end = month_start_date, month_end_date
        prev_period_start, prev_period_end = prev_month_start, prev_month_end

    cache_scope = "all" if can_view_all else f"prof-{prof.id if prof else request.user.id}"
    cache_key = f"dashboard:metrics:v3:{cache_scope}:{today.isoformat()}:{period}"
    metrics = cache.get(cache_key)
    if metrics is None:
        # Cache only counters/charts for a short period. Agenda list stays uncached.
        revenue_totals = revenue_qs.aggregate(
            revenue_today=Coalesce(
                Sum("final_price", filter=Q(date__gte=today_start, date__lt=today_end)),
                Decimal("0.00"),
            ),
            revenue_yesterday=Coalesce(
                Sum("final_price", filter=Q(date__gte=yesterday_start, date__lt=yesterday_end)),
                Decimal("0.00"),
            ),
            revenue_period=Coalesce(
                Sum("final_price", filter=Q(date__gte=period_start, date__lt=period_end)),
                Decimal("0.00"),
            ),
            revenue_prev_period=Coalesce(
                Sum("final_price", filter=Q(date__gte=prev_period_start, date__lt=prev_period_end)),
                Decimal("0.00"),
            ),
        )
        revenue_today = revenue_totals["revenue_today"] or Decimal("0.00")
        revenue_yesterday = revenue_totals["revenue_yesterday"] or Decimal("0.00")
        revenue_period = revenue_totals["revenue_period"] or Decimal("0.00")
        revenue_prev_period = revenue_totals["revenue_prev_period"] or Decimal("0.00")

        appt_totals = appt_metrics_qs.aggregate(
            appointments_today=Count("id", filter=Q(date=today)),
            appointments_yesterday=Count("id", filter=Q(date=today - timedelta(days=1))),
            appointments_period=Count("id", filter=Q(date__gte=period_start, date__lt=period_end)),
            debt_today_count=Count(
                "id",
                filter=Q(date=today, status=Appointment.STATUS_IN_DEBT),
            ),
            debt_today_amount=Coalesce(
                Sum("final_price", filter=Q(date=today, status=Appointment.STATUS_IN_DEBT)),
                Decimal("0.00"),
            ),
            cancelled_today_count=Count(
                "id",
                filter=Q(date=today, status=Appointment.STATUS_CANCELLED),
            ),
            completed_today_count=Count(
                "id",
                filter=Q(date=today, status=Appointment.STATUS_COMPLETED),
            ),
            pending_today_count=Count(
                "id",
                filter=Q(date=today, status=Appointment.STATUS_PENDING),
            ),
            scheduled_count=Count(
                "id",
                filter=Q(
                    date__gte=period_start,
                    date__lt=period_end,
                    status=Appointment.STATUS_SCHEDULED,
                ),
            ),
            completed_count=Count(
                "id",
                filter=Q(
                    date__gte=period_start,
                    date__lt=period_end,
                    status=Appointment.STATUS_COMPLETED,
                ),
            ),
            cancelled_count=Count(
                "id",
                filter=Q(
                    date__gte=period_start,
                    date__lt=period_end,
                    status=Appointment.STATUS_CANCELLED,
                ),
            ),
            prev_total=Count("id", filter=Q(date__gte=prev_period_start, date__lt=prev_period_end)),
            prev_completed=Count(
                "id",
                filter=Q(
                    date__gte=prev_period_start,
                    date__lt=prev_period_end,
                    status=Appointment.STATUS_COMPLETED,
                ),
            ),
            weekly_appointments_count=Count(
                "id",
                filter=Q(date__gte=week_start_date, date__lt=week_end_date)
                & ~Q(status=Appointment.STATUS_CANCELLED),
            ),
            monthly_appointments_count=Count(
                "id",
                filter=Q(date__gte=month_start_date, date__lt=month_end_date)
                & ~Q(status=Appointment.STATUS_CANCELLED),
            ),
        )
        appointments_today = appt_totals["appointments_today"] or 0
        appointments_yesterday = appt_totals["appointments_yesterday"] or 0
        appointments_period = appt_totals["appointments_period"] or 0
        debt_today_count = appt_totals["debt_today_count"] or 0
        debt_today_amount = appt_totals["debt_today_amount"] or Decimal("0.00")
        cancelled_today_count = appt_totals["cancelled_today_count"] or 0
        completed_today_count = appt_totals["completed_today_count"] or 0
        pending_today_count = appt_totals["pending_today_count"] or 0
        scheduled_count = appt_totals["scheduled_count"] or 0
        completed_count = appt_totals["completed_count"] or 0
        cancelled_count = appt_totals["cancelled_count"] or 0
        prev_total = appt_totals["prev_total"] or 0
        prev_completed = appt_totals["prev_completed"] or 0
        weekly_appointments_count = appt_totals["weekly_appointments_count"] or 0
        monthly_appointments_count = appt_totals["monthly_appointments_count"] or 0

        def _rate(completed: int, total: int) -> Decimal:
            if not total:
                return Decimal("0.0")
            return (Decimal(completed) / Decimal(total)) * Decimal("100")

        completed_rate = _rate(completed_count, appointments_period)
        prev_completed_rate = _rate(prev_completed, prev_total)
        if prev_completed_rate > 0:
            completed_rate_delta = ((completed_rate - prev_completed_rate) / prev_completed_rate) * Decimal("100")
            completed_rate_delta_display = f"{completed_rate_delta:.1f}%"
            if completed_rate_delta > 0:
                completed_rate_delta_display = f"+{completed_rate_delta_display}"
        else:
            completed_rate_delta_display = "—"

        if period == "day":
            delta_label = "dia anterior"
        elif period == "week":
            delta_label = "semana anterior"
        else:
            delta_label = "mês anterior"

        total_pct = 100
        scheduled_pct = round((scheduled_count / appointments_period) * 100) if appointments_period else 0
        completed_pct = round((completed_count / appointments_period) * 100) if appointments_period else 0
        cancelled_pct = round((cancelled_count / appointments_period) * 100) if appointments_period else 0

        chart_labels = []
        chart_values = []
        if period == "day":
            hourly = {h: 0 for h in range(24)}
            for time_value in appt_metrics_qs.filter(date=today).values_list("time", flat=True):
                if time_value:
                    hourly[time_value.hour] += 1
            for hour in range(24):
                chart_labels.append(f"{hour:02d}:00")
                chart_values.append(hourly[hour])
        else:
            # Single grouped query avoids daily count loops.
            daily_counts = dict(
                appt_metrics_qs
                .filter(date__gte=period_start, date__lt=period_end)
                .values("date")
                .annotate(total=Count("id"))
                .values_list("date", "total")
            )
            day = period_start
            while day < period_end:
                chart_labels.append(day.strftime("%d %b"))
                chart_values.append(daily_counts.get(day, 0))
                day += timedelta(days=1)

        client_qs = ClientProfile.objects.all()
        if prof:
            client_qs = client_qs.filter(appointments__professional=prof).distinct()
        period_start_dt = timezone.make_aware(datetime.combine(period_start, datetime.min.time()))
        period_end_dt = timezone.make_aware(datetime.combine(period_end, datetime.min.time()))
        prev_start_dt = timezone.make_aware(datetime.combine(prev_period_start, datetime.min.time()))
        prev_end_dt = timezone.make_aware(datetime.combine(prev_period_end, datetime.min.time()))
        today_start_dt = timezone.make_aware(datetime.combine(today, datetime.min.time()))
        today_end_dt = timezone.make_aware(datetime.combine(today + timedelta(days=1), datetime.min.time()))

        client_totals = client_qs.aggregate(
            new_clients_today=Count("id", filter=Q(created_at__gte=today_start_dt, created_at__lt=today_end_dt)),
            new_clients_period=Count("id", filter=Q(created_at__gte=period_start_dt, created_at__lt=period_end_dt)),
            new_clients_prev=Count("id", filter=Q(created_at__gte=prev_start_dt, created_at__lt=prev_end_dt)),
        )
        new_clients_today = client_totals["new_clients_today"] or 0
        new_clients_period = client_totals["new_clients_period"] or 0
        new_clients_prev = client_totals["new_clients_prev"] or 0

        new_clients_trend = compute_trend(Decimal(new_clients_period), Decimal(new_clients_prev))
        revenue_trend_today = compute_trend(revenue_today, revenue_yesterday)
        revenue_trend_period = compute_trend(revenue_period, revenue_prev_period)
        appointments_trend = compute_trend(Decimal(appointments_today), Decimal(appointments_yesterday))

        appointments_by_service = list(
            appt_metrics_qs
            .filter(date__gte=period_start, date__lt=period_end)
            .exclude(status=Appointment.STATUS_CANCELLED)
            .values("service__name")
            .annotate(total=Count("id"))
            .order_by("-total")
        )

        weekly_counts_by_date = dict(
            appt_metrics_qs
            .filter(date__gte=week_start_date, date__lt=week_end_date)
            .exclude(status=Appointment.STATUS_CANCELLED)
            .values("date")
            .annotate(total=Count("id"))
            .values_list("date", "total")
        )
        week_days = [week_start_date + timedelta(days=i) for i in range(7)]
        weekly_chart_series = [weekly_counts_by_date.get(day, 0) for day in week_days]

        month_counts_by_date = dict(
            appt_metrics_qs
            .filter(date__gte=month_start_date, date__lt=month_end_date)
            .exclude(status=Appointment.STATUS_CANCELLED)
            .values("date")
            .annotate(total=Count("id"))
            .values_list("date", "total")
        )
        month_weeks = []
        current = month_start_date
        while current < month_end_date:
            week_end = min(current + timedelta(days=7), month_end_date)
            bucket_total = sum(
                count
                for day, count in month_counts_by_date.items()
                if current <= day < week_end
            )
            month_weeks.append(bucket_total)
            current = week_end

        stock_alert_count = (
            Product.objects
            .filter(is_active=True)
            .annotate(
                stock=Coalesce(
                    Sum("movements__quantity_base", filter=Q(movements__is_void=False)),
                    Decimal("0.00"),
                )
            )
            .filter(stock__lte=F("min_stock_alert"))
            .count()
        )

        metrics = {
            "revenue_today": revenue_today,
            "revenue_trend_today": revenue_trend_today,
            "appointments_today": appointments_today,
            "appointments_trend": appointments_trend,
            "new_clients_period": new_clients_period,
            "new_clients_trend": new_clients_trend,
            "new_clients_today": new_clients_today,
            "revenue_period": revenue_period,
            "revenue_trend_period": revenue_trend_period,
            "appointments_period": appointments_period,
            "appointments_by_service": appointments_by_service,
            "scheduled_count": scheduled_count,
            "completed_count": completed_count,
            "cancelled_count": cancelled_count,
            "debt_today_count": debt_today_count,
            "debt_today_amount": debt_today_amount,
            "cancelled_today_count": cancelled_today_count,
            "completed_today_count": completed_today_count,
            "pending_today_count": pending_today_count,
            "completed_rate_delta_display": completed_rate_delta_display,
            "delta_label": delta_label,
            "total_pct": total_pct,
            "scheduled_pct": scheduled_pct,
            "completed_pct": completed_pct,
            "cancelled_pct": cancelled_pct,
            "chart_labels": chart_labels,
            "chart_values": chart_values,
            "weekly_appointments_count": weekly_appointments_count,
            "monthly_appointments_count": monthly_appointments_count,
            "stock_alert_count": stock_alert_count,
            "weekly_chart_series": weekly_chart_series,
            "monthly_chart_series": month_weeks,
        }
        cache.set(cache_key, metrics, 120)

    agenda_qs = agenda_base_qs.filter(date=today).order_by("time")
    agenda_page_size = 10
    agenda_page_obj = Paginator(agenda_qs, agenda_page_size).get_page(request.GET.get("agenda_page") or 1)
    agenda_today = []
    for appt in agenda_page_obj.object_list:
        agenda_today.append(
            {
                "id": appt.id,
                "time": appt.time,
                "client": appt.client.get_full_name() or appt.client.username,
                "service": appt.service.name if appt.service else "-",
                "professional": appt.professional.user.get_full_name() or appt.professional.user.username,
                "status": appt.get_status_display(),
                "status_value": appt.status,
                "open_url": reverse("professional_appointment_detail", args=[appt.id]),
            }
        )

    return render(
        request,
        "core/backoffice/dashboard.html",
        {
            "today": today,
            "is_admin": can_view_all,
            "revenue_today": metrics["revenue_today"],
            "revenue_trend_today": metrics["revenue_trend_today"],
            "appointments_today": metrics["appointments_today"],
            "appointments_trend": metrics["appointments_trend"],
            "new_clients_period": metrics["new_clients_period"],
            "new_clients_trend": metrics["new_clients_trend"],
            "revenue_period": metrics["revenue_period"],
            "revenue_trend_period": metrics["revenue_trend_period"],
            "appointments_period": metrics["appointments_period"],
            "period": period,
            "agenda_today": agenda_today,
            "agenda_page_obj": agenda_page_obj,
            "appointments_by_service": metrics["appointments_by_service"],
            "scheduled_count": metrics["scheduled_count"],
            "completed_count": metrics["completed_count"],
            "cancelled_count": metrics["cancelled_count"],
            "debt_today_count": metrics["debt_today_count"],
            "debt_today_amount": metrics["debt_today_amount"],
            "cancelled_today_count": metrics["cancelled_today_count"],
            "completed_today_count": metrics["completed_today_count"],
            "pending_today_count": metrics["pending_today_count"],
            "completed_rate_delta_display": metrics["completed_rate_delta_display"],
            "delta_label": metrics["delta_label"],
            "total_pct": metrics["total_pct"],
            "scheduled_pct": metrics["scheduled_pct"],
            "completed_pct": metrics["completed_pct"],
            "cancelled_pct": metrics["cancelled_pct"],
            "chart_labels": metrics["chart_labels"],
            "chart_values": metrics["chart_values"],
            "kpi_revenue_today": metrics["revenue_today"],
            "kpi_appts_today": metrics["appointments_today"],
            "kpi_new_clients": metrics["new_clients_today"],
            "kpi_sales_total": metrics["revenue_period"],
            "can_view_financial": can_view_financial,
            "today_agenda": agenda_today,
            "weekly_appointments_count": metrics["weekly_appointments_count"],
            "monthly_appointments_count": metrics["monthly_appointments_count"],
            "stock_alert_count": metrics["stock_alert_count"],
            "weekly_chart_series": json.dumps(metrics["weekly_chart_series"]),
            "monthly_chart_series": json.dumps(metrics["monthly_chart_series"]),
        },
    )


def _agenda_export_selected_csv_response(appointments):
    buffer = io.StringIO()
    writer = csv.writer(buffer, delimiter=";")
    headers = [
        "Referencia",
        "Data",
        "Hora",
        "Servico",
        "Profissional",
        "Cliente",
        "Telefone",
        "NIF",
        "Estado",
        "Pagamento",
        "Preco",
        "Parceria",
    ]
    writer.writerow(headers)

    for appt in appointments:
        client_profile = getattr(appt.client, "client_profile", None)
        client_name = (
            (client_profile.full_name if client_profile and client_profile.full_name else "")
            or appt.client.get_full_name()
            or appt.client.username
        )
        payment_status = "Pago" if appt.is_paid else ("Em divida" if appt.status == Appointment.STATUS_IN_DEBT else "Por pagar")
        writer.writerow(
            [
                f"MC-{appt.id:06d}",
                appt.date.strftime("%d/%m/%Y") if appt.date else "",
                appt.time.strftime("%H:%M") if appt.time else "",
                appt.service.name if appt.service else "-",
                appt.professional.user.get_full_name() or appt.professional.user.username,
                client_name,
                getattr(client_profile, "phone", "") if client_profile else "",
                getattr(client_profile, "nif", "") if client_profile else "",
                appt.get_status_display(),
                payment_status,
                f"{Decimal(appt.final_price or 0):.2f}",
                appt.partner.name if appt.partner else "",
            ]
        )

    payload = buffer.getvalue()
    response = HttpResponse(payload, content_type="text/csv; charset=utf-8")
    filename = timezone.localtime().strftime("agenda_marcacoes_%Y%m%d_%H%M.csv")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _agenda_export_selected_excel_response(appointments):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except Exception:
        return _agenda_export_selected_csv_response(appointments)

    wb = Workbook()
    ws = wb.active
    ws.title = "Marcacoes"

    headers = [
        "Referencia",
        "Data",
        "Hora",
        "Servico",
        "Profissional",
        "Cliente",
        "Telefone",
        "NIF",
        "Estado",
        "Pagamento",
        "Preco",
        "Parceria",
    ]
    ws.append(headers)
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.font = Font(bold=True)

    for appt in appointments:
        client_profile = getattr(appt.client, "client_profile", None)
        client_name = (
            (client_profile.full_name if client_profile and client_profile.full_name else "")
            or appt.client.get_full_name()
            or appt.client.username
        )
        payment_status = "Pago" if appt.is_paid else ("Em divida" if appt.status == Appointment.STATUS_IN_DEBT else "Por pagar")
        ws.append(
            [
                f"MC-{appt.id:06d}",
                appt.date.strftime("%d/%m/%Y") if appt.date else "",
                appt.time.strftime("%H:%M") if appt.time else "",
                appt.service.name if appt.service else "-",
                appt.professional.user.get_full_name() or appt.professional.user.username,
                client_name,
                getattr(client_profile, "phone", "") if client_profile else "",
                getattr(client_profile, "nif", "") if client_profile else "",
                appt.get_status_display(),
                payment_status,
                float(Decimal(appt.final_price or 0)),
                appt.partner.name if appt.partner else "",
            ]
        )

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 28
    ws.column_dimensions["F"].width = 28
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 20
    ws.column_dimensions["J"].width = 14
    ws.column_dimensions["K"].width = 12
    ws.column_dimensions["L"].width = 22

    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=11).number_format = '#,##0.00 "EUR"'

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = timezone.localtime().strftime("agenda_marcacoes_%Y%m%d_%H%M.xlsx")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def backoffice_agenda_view(request):
    if not request.user.is_authenticated:
        return redirect(f"{reverse('login')}?next={request.get_full_path()}")

    view_all = can_view_all_calendar(request.user)
    prof = Professional.objects.filter(user=request.user).first()
    if not view_all and not prof:
        return HttpResponseForbidden("Acesso apenas para profissionais.")
    is_technician_agenda = not view_all

    today = timezone.localdate()
    now_t = timezone.localtime().time()
    quick_modal_open = False
    quick_errors = []
    quick_slots = []
    quick_form = {
        "client_id": "",
        "client_label": "",
        "service_id": "",
        "professional_id": "",
        "date": "",
        "time": "",
        "symptomatology": "",
    }

    if request.method == "POST":
        post_action = (request.POST.get("action") or "").strip()
        if post_action == "bulk_update":
            return_to = _safe_return_to(request, request.POST.get("return_to")) or reverse("backoffice_agenda")
            bulk_action = (request.POST.get("bulk_action") or "").strip()
            selected_ids = []
            for raw in request.POST.getlist("appointment_ids"):
                try:
                    selected_ids.append(int(raw))
                except (TypeError, ValueError):
                    continue

            valid_actions = {
                "confirm_selected",
                "confirm_and_paid_selected",
                "mark_no_show_selected",
                "mark_completed_selected",
                "mark_in_debt_selected",
                "export_selected",
            }
            if bulk_action not in valid_actions:
                messages.error(request, "Ação inválida.")
                return redirect(return_to)
            if not selected_ids:
                messages.error(request, "Seleciona pelo menos uma marcação.")
                return redirect(return_to)

            selected_qs = Appointment.objects.select_related(
                "client",
                "client__client_profile",
                "service",
                "professional",
                "professional__user",
                "partner",
            ).filter(id__in=selected_ids)
            if not view_all and prof:
                selected_qs = selected_qs.filter(professional=prof)

            selected_qs = selected_qs.order_by("date", "time", "id")

            if bulk_action == "export_selected":
                return _agenda_export_selected_excel_response(selected_qs)

            result = apply_bulk_appointment_action(
                appointments=selected_qs,
                action=bulk_action,
                actor=request.user,
                today=today,
                now_t=now_t,
                audit_source="agenda",
            )
            for transition in result["status_transitions"]:
                sync_subcontractor_payout(transition["appointment"], actor=request.user)
            skipped_note = ""
            if result["skipped_future"] or result["skipped_locked"] or result["skipped_unpaid"]:
                skipped_note = (
                    " ("
                    f"{result['skipped_future']} no futuro, "
                    f"{result['skipped_locked']} em estados fechados, "
                    f"{result['skipped_unpaid']} sem pagamento ignoradas"
                    ")"
                )
            if bulk_action == "confirm_and_paid_selected":
                messages.success(
                    request,
                    (
                        f"Atualizadas {result['status_changed']} marcações e marcadas "
                        f"{result['paid_changed']} como pagas{skipped_note}."
                    ),
                )
            elif bulk_action == "mark_no_show_selected":
                messages.success(request, f"Marcadas {result['status_changed']} marcações como falta{skipped_note}.")
            elif bulk_action == "mark_completed_selected":
                messages.success(request, f"Marcadas {result['status_changed']} marcações como concluídas{skipped_note}.")
            elif bulk_action == "mark_in_debt_selected":
                messages.success(request, f"Marcadas {result['status_changed']} marcações como em dívida{skipped_note}.")
            else:
                messages.success(request, f"Atualizadas {result['status_changed']} marcações{skipped_note}.")
            return redirect(return_to)

    if request.method == "POST" and request.POST.get("action") == "quick_create":
        if not can_book_for_any_professional(request.user):
            return HttpResponseForbidden("Apenas receção/admin pode criar marcações aqui.")

        quick_form = {
            "client_id": (request.POST.get("client_id") or "").strip(),
            "service_id": (request.POST.get("service_id") or "").strip(),
            "professional_id": (request.POST.get("professional_id") or "").strip(),
            "date": (request.POST.get("date") or "").strip(),
            "time": (request.POST.get("time") or "").strip(),
            "symptomatology": (request.POST.get("symptomatology") or "").strip(),
            "client_label": (request.POST.get("client_label") or "").strip(),
        }

        return_to = _safe_return_to(request, request.POST.get("return_to")) or reverse("backoffice_agenda")

        if not quick_form["client_id"]:
            quick_errors.append("Seleciona um cliente.")
        if not quick_form["service_id"]:
            quick_errors.append("Seleciona um serviço.")
        if not quick_form["professional_id"]:
            quick_errors.append("Seleciona um profissional.")
        if not quick_form["date"]:
            quick_errors.append("Seleciona uma data.")
        if not quick_form["time"]:
            quick_errors.append("Seleciona uma hora.")

        client_profile = None
        client_user = None
        if quick_form["client_id"]:
            client_profile = ClientProfile.objects.select_related("user").filter(id=quick_form["client_id"]).first()
            if not client_profile:
                quick_errors.append("Cliente inválido.")
            else:
                client_user = client_profile.user
                quick_form["client_label"] = client_profile.full_name or (client_user.get_full_name() if client_user else "") or str(client_profile)
                if not client_user:
                    quick_errors.append("Este cliente ainda não tem utilizador associado.")

        service = None
        prof = None
        if quick_form["service_id"]:
            service = Service.objects.filter(id=quick_form["service_id"]).first()
            if not service:
                quick_errors.append("Serviço inválido.")

        if quick_form["professional_id"]:
            prof = Professional.objects.filter(id=quick_form["professional_id"]).first()
            if not prof:
                quick_errors.append("Profissional inválido.")

        date_obj = None
        time_obj = None
        if quick_form["date"]:
            try:
                date_obj = datetime.strptime(quick_form["date"], "%Y-%m-%d").date()
            except Exception:
                quick_errors.append("Data inválida.")
        if quick_form["time"]:
            try:
                time_obj = datetime.strptime(quick_form["time"], "%H:%M").time()
            except Exception:
                quick_errors.append("Hora inválida.")

        if service and prof:
            if not Professional.objects.filter(id=prof.id, services__id=service.id).exists():
                quick_errors.append("Profissional inválido para este serviço.")

        if service and prof and date_obj:
            if date_obj < today:
                quick_errors.append("Não podes marcar no passado.")
            elif date_obj == today and time_obj and time_obj <= timezone.localtime().time():
                quick_errors.append("Este horário já passou.")
            elif not professional_works_on_date(prof, date_obj):
                prof_days = professional_weekdays_labels(prof)
                quick_errors.append(
                    f"Este profissional não atende nesse dia. Atende: {', '.join(prof_days) or '—'}."
                )

        if service and prof and date_obj:
            quick_slots = _get_slots(prof, date_obj, service=service)

        if service and prof and date_obj and time_obj and not quick_errors:
            if _is_slot_blocked(prof, date_obj, time_obj):
                quick_errors.append("Este horário está indisponível.")
            elif quick_form["time"] not in quick_slots:
                quick_errors.append("Esse horário já não está disponível.")

        if quick_errors:
            quick_modal_open = True
        else:
            try:
                with transaction.atomic():
                    pricing = compute_pricing(service, client_profile)
                    reactivated_cancelled = _find_matching_cancelled_appointment(
                        client_user=client_user,
                        professional=prof,
                        service=service,
                        date_obj=date_obj,
                        time_obj=time_obj,
                    )
                    if reactivated_cancelled:
                        old_status = reactivated_cancelled.status
                        reactivated_cancelled.symptomatology = quick_form["symptomatology"]
                        reactivated_cancelled.status = Appointment.STATUS_SCHEDULED
                        reactivated_cancelled.base_price = pricing["base_price_applied"]
                        reactivated_cancelled.partner = pricing["partner"]
                        reactivated_cancelled.partner_price = pricing["partner_price_applied"]
                        reactivated_cancelled.discount_type = pricing["discount_type"]
                        reactivated_cancelled.discount_value = pricing["discount_value"]
                        reactivated_cancelled.final_price = pricing["final_price"]
                        reactivated_cancelled.session_index = pricing["session_index"]
                        reactivated_cancelled.pricing_tier = pricing["pricing_tier"]
                        reactivated_cancelled.base_price_applied = pricing["base_price_applied"]
                        reactivated_cancelled.partner_price_applied = pricing["partner_price_applied"]
                        reactivated_cancelled.discount_applied = pricing["discount_applied"]
                        reactivated_cancelled.is_paid = False
                        reactivated_cancelled.paid_at = None
                        reactivated_cancelled.completed_by = None
                        reactivated_cancelled.completed_at = None
                        reactivated_cancelled.save(
                            update_fields=[
                                "symptomatology",
                                "status",
                                "base_price",
                                "partner",
                                "partner_price",
                                "discount_type",
                                "discount_value",
                                "final_price",
                                "session_index",
                                "pricing_tier",
                                "base_price_applied",
                                "partner_price_applied",
                                "discount_applied",
                                "is_paid",
                                "paid_at",
                                "completed_by",
                                "completed_at",
                            ]
                        )
                        appt = reactivated_cancelled
                        log_appt(
                            AppointmentLog.ACTION_STATUS_UPDATED,
                            appt,
                            request.user,
                            old_status=old_status,
                            new_status=appt.status,
                            note="Reativada no backoffice a partir de uma marcação cancelada do mesmo slot.",
                            request=request,
                        )
                        success_message = "Marcação cancelada anterior reativada com sucesso."
                    else:
                        appt = Appointment.objects.create(
                            client=client_user,
                            professional=prof,
                            service=service,
                            date=date_obj,
                            time=time_obj,
                            symptomatology=quick_form["symptomatology"],
                            base_price=pricing["base_price_applied"],
                            partner=pricing["partner"],
                            partner_price=pricing["partner_price_applied"],
                            discount_type=pricing["discount_type"],
                            discount_value=pricing["discount_value"],
                            final_price=pricing["final_price"],
                            session_index=pricing["session_index"],
                            pricing_tier=pricing["pricing_tier"],
                            base_price_applied=pricing["base_price_applied"],
                            partner_price_applied=pricing["partner_price_applied"],
                            discount_applied=pricing["discount_applied"],
                        )
                        log_appt(
                            AppointmentLog.ACTION_CREATED,
                            appt,
                            request.user,
                            note="Criada no backoffice",
                            request=request,
                        )
                        success_message = "Marcação criada com sucesso."
                messages.success(request, success_message)
                return redirect(return_to)
            except IntegrityError:
                quick_errors.append("Esse horário já não está disponível.")
                quick_modal_open = True
    date_str = (request.GET.get("date") or "").strip()
    tab = (request.GET.get("tab") or "all").strip()
    if tab not in ("all", "pending", "review"):
        tab = "all"
    view_mode = (request.GET.get("view_mode") or "day").strip()
    if view_mode not in ("day", "week", "all"):
        view_mode = "day"
    professional_id = (request.GET.get("professional_id") or "").strip()
    service_id = (request.GET.get("service_id") or "").strip()
    status = (request.GET.get("status") or "").strip()
    kind = (request.GET.get("type") or "all").strip()
    q = (request.GET.get("q") or "").strip()
    client_id = (request.GET.get("client_id") or "").strip()
    per_page = request.GET.get("per_page") or "10"

    try:
        per_page = int(per_page)
    except (TypeError, ValueError):
        per_page = 10
    if per_page not in (10, 25, 50):
        per_page = 10

    try:
        date_obj = datetime.strptime(date_str, "%Y-%m-%d").date() if date_str else today
    except Exception:
        date_obj = today

    if view_mode == "week":
        start_date = date_obj - timedelta(days=date_obj.weekday())
        end_date = start_date + timedelta(days=6)
    elif view_mode == "all":
        start_date = None
        end_date = None
    else:
        start_date = date_obj
        end_date = date_obj

    if not view_all:
        professional_id = str(prof.id)
        kind = "appointment"

    appointments = Appointment.objects.select_related(
        "client",
        "client__client_profile",
        "service",
        "professional",
        "professional__user",
    ).prefetch_related(
        "consumptions__product",
    )
    if start_date and end_date:
        appointments = appointments.filter(date__range=(start_date, end_date))

    if professional_id:
        appointments = appointments.filter(professional_id=professional_id)
    if service_id:
        appointments = appointments.filter(service_id=service_id)
    if client_id:
        appointments = appointments.filter(client__client_profile__id=client_id)
    if q:
        def _extract_appointment_reference_id(value):
            text = str(value or "").strip()
            if not text:
                return None

            candidates = [text] + text.split()
            for candidate in candidates:
                token = candidate.strip().upper().replace("#", "")
                compact = "".join(ch for ch in token if ch.isalnum())
                if not compact:
                    continue
                if compact.startswith("MC"):
                    digits = compact[2:]
                else:
                    digits = compact
                if digits.isdigit():
                    try:
                        return int(digits)
                    except Exception:
                        continue
            return None

        appointments_by_text = apply_terms_filter(
            appointments,
            q,
            [
                "client__username__icontains",
                "client__first_name__icontains",
                "client__last_name__icontains",
                "client__client_profile__full_name__icontains",
                "client__client_profile__phone__icontains",
                "client__client_profile__nif__icontains",
            ],
        )
        appointment_ref_id = _extract_appointment_reference_id(q)
        if appointment_ref_id:
            appointments = appointments.filter(
                Q(id=appointment_ref_id) | Q(id__in=appointments_by_text.values("id"))
            )
        else:
            appointments = appointments_by_text

    appointments_base = appointments

    group_sessions = GroupSession.objects.select_related(
        "service",
        "professional",
        "professional__user",
    )
    if start_date and end_date:
        group_sessions = group_sessions.filter(date__range=(start_date, end_date))

    if professional_id:
        group_sessions = group_sessions.filter(professional_id=professional_id)
    if service_id:
        group_sessions = group_sessions.filter(service_id=service_id)

    group_sessions_base = group_sessions

    if status and tab == "all":
        appointments = appointments.filter(status=status)
        group_sessions = group_sessions.filter(status=status)

    if tab == "pending":
        appointments = appointments.filter(
            status=Appointment.STATUS_PENDING,
        ).filter(
            Q(date__gt=today) | Q(date=today, time__gte=now_t)
        )
        kind = "appointment"
        status = ""
        group_sessions = GroupSession.objects.none()
    elif tab == "review":
        appointments = appointments.exclude(
            status__in=[
                Appointment.STATUS_CANCELLED,
                Appointment.STATUS_IN_DEBT,
                Appointment.STATUS_COMPLETED,
                Appointment.STATUS_NO_SHOW,
            ]
        ).filter(
            Q(date__lt=today) | Q(date=today, time__lt=now_t)
        ).filter(
            status__in=[
                Appointment.STATUS_PENDING,
                Appointment.STATUS_SCHEDULED,
                Appointment.STATUS_AWAITING_VALIDATION,
            ]
        )
        kind = "appointment"
        status = ""
        group_sessions = GroupSession.objects.none()

    # Counts for tabs (always based on appointment filters)
    pending_count = appointments_base.filter(
        status=Appointment.STATUS_PENDING,
    ).filter(
        Q(date__gt=today) | Q(date=today, time__gte=now_t)
    ).count()
    review_count = appointments_base.exclude(
        status__in=[
            Appointment.STATUS_CANCELLED,
            Appointment.STATUS_IN_DEBT,
            Appointment.STATUS_COMPLETED,
            Appointment.STATUS_NO_SHOW,
        ]
    ).filter(
        Q(date__lt=today) | Q(date=today, time__lt=now_t)
    ).filter(
        status__in=[
            Appointment.STATUS_PENDING,
            Appointment.STATUS_SCHEDULED,
            Appointment.STATUS_AWAITING_VALIDATION,
        ]
    ).count()
    if kind == "appointment":
        all_count = appointments.count()
    elif kind == "group":
        all_count = group_sessions.count()
    else:
        all_count = appointments.count() + group_sessions.count()

    if kind == "appointment":
        group_sessions = GroupSession.objects.none()
    elif kind == "group":
        appointments = Appointment.objects.none()

    group_sessions = group_sessions.annotate(
        active_count=Count("enrolments", filter=Q(enrolments__status__in=group_booked_statuses()))
    )

    return_to = request.get_full_path()
    items = []
    def format_decimal(value):
        try:
            raw = f"{Decimal(value):f}"
        except Exception:
            return str(value)
        raw = raw.rstrip("0").rstrip(".")
        return raw if raw else "0"

    for appt in appointments:
        client_profile = getattr(appt.client, "client_profile", None)
        client_name = (
            (client_profile.full_name if client_profile and client_profile.full_name else "")
            or appt.client.get_full_name()
            or appt.client.username
        )
        consumptions = list(appt.consumptions.all())
        if consumptions:
            parts = []
            for cons in consumptions:
                product = cons.product
                if not product:
                    continue
                unit = product.get_unit_base_display()
                qty = format_decimal(cons.quantity_base)
                parts.append(f"{product.name} ({qty} {unit})")
            consumptions_label = ", ".join(parts) if parts else "—"
        else:
            consumptions_label = "—"
        if view_all:
            cancel_url = f"{reverse('backoffice_cancel_appointment', args=[appt.id])}?return_to={urlencode({'return_to': return_to})[10:]}"
            complete_url = f"{reverse('backoffice_complete_appointment', args=[appt.id])}?return_to={urlencode({'return_to': return_to})[10:]}"
            reschedule_url = f"{reverse('reschedule_appointment', args=[appt.id])}?return_to={urlencode({'return_to': return_to})[10:]}"
        else:
            cancel_url = f"{reverse('professional_cancel_appointment', args=[appt.id])}?return_to={urlencode({'return_to': return_to})[10:]}"
            complete_url = f"{reverse('professional_complete_appointment', args=[appt.id])}?return_to={urlencode({'return_to': return_to})[10:]}"
            reschedule_url = f"{reverse('professional_reschedule_appointment', args=[appt.id])}?return_to={urlencode({'return_to': return_to})[10:]}"
        items.append(
            AgendaItem(
                kind="appointment",
                appointment_id=appt.id,
                date=appt.date,
                time=appt.time,
                service_name=appt.service.name if appt.service else "-",
                professional_name=appt.professional.user.get_full_name() or appt.professional.user.username,
                client_label=client_name,
                consumptions_label=consumptions_label,
                status_label=appt.get_status_display(),
                status_raw=appt.status,
                price_label=str(appt.final_price) if appt.final_price is not None else "—",
                open_url=f"{reverse('professional_appointment_detail', args=[appt.id])}?return_to={urlencode({'return_to': return_to})[10:]}",
                cancel_url=cancel_url,
                complete_url=complete_url,
                reschedule_url=reschedule_url,
            )
        )

    for session in group_sessions:
        professional_name = "—"
        if session.professional:
            professional_name = session.professional.user.get_full_name() or session.professional.user.username
        items.append(
            AgendaItem(
                kind="group",
                appointment_id=None,
                date=session.date,
                time=session.time,
                service_name=session.service.name if session.service else "-",
                professional_name=professional_name,
                client_label=f"{session.active_count}/{session.capacity_value} inscritos",
                consumptions_label="—",
                status_label=session.get_status_display(),
                status_raw=session.status,
                price_label="—",
                open_url=reverse("group_session_detail_admin", args=[session.id]),
                cancel_url=None,
                complete_url=None,
                reschedule_url=None,
            )
        )

    items.sort(key=lambda item: (item.date, item.time, item.kind))
    paginator = Paginator(items, per_page)
    page_obj = paginator.get_page(request.GET.get("page") or 1)
    has_bulk_candidates = any(item.kind == "appointment" and item.appointment_id for item in page_obj.object_list)

    return render(
        request,
        "backoffice/agenda.html",
        {
            "items": page_obj.object_list,
            "page_obj": page_obj,
            "paginator": paginator,
            "today": today,
            "date": date_obj,
            "view_mode": view_mode,
            "professional_id": professional_id,
            "service_id": service_id,
            "status": status,
            "tab": tab,
            "type": kind,
            "tab_counts": {
                "all": all_count,
                "pending": pending_count,
                "review": review_count,
            },
            "q": q,
            "per_page": per_page,
            "professionals": Professional.objects.select_related("user").order_by("user__username") if view_all else [prof],
            "services": Service.objects.order_by("name"),
            "return_to": return_to,
            "can_quick_create": can_book_for_any_professional(request.user),
            "show_professional_filter": view_all,
            "show_type_filter": view_all,
            "is_technician_agenda": is_technician_agenda,
            "quick_modal_open": quick_modal_open,
            "quick_errors": quick_errors,
            "quick_form": quick_form,
            "quick_slots": quick_slots,
            "has_bulk_candidates": has_bulk_candidates,
        },
    )


def backoffice_faturacao_view(request):
    if not is_admin_role(request.user):
        return HttpResponseForbidden("Acesso reservado a administradores.")

    today = timezone.localdate()
    per_page = request.GET.get("per_page") or "10"
    try:
        per_page = int(per_page)
    except (TypeError, ValueError):
        per_page = 10
    if per_page not in (5, 10, 15, 25, 50):
        per_page = 10

    start_param = (request.GET.get("start") or "").strip()
    end_param = (request.GET.get("end") or "").strip()
    professional_id = (request.GET.get("professional_id") or "").strip()
    service_id = (request.GET.get("service_id") or "").strip()

    if start_param:
        try:
            start_date = datetime.strptime(start_param, "%Y-%m-%d").date()
        except ValueError:
            start_date = today - timedelta(days=29)
    else:
        start_date = today - timedelta(days=29)

    if end_param:
        try:
            end_date = datetime.strptime(end_param, "%Y-%m-%d").date()
        except ValueError:
            end_date = today
    else:
        end_date = today

    revenue_qs = get_revenue_queryset(request.user)
    debt_qs = Appointment.objects.filter(status=Appointment.STATUS_IN_DEBT, date__lte=today)
    group_revenue_qs = GroupMonthlyCharge.objects.filter(status=GroupMonthlyCharge.STATUS_PAID)
    group_debt_qs = GroupMonthlyCharge.objects.filter(status=GroupMonthlyCharge.STATUS_UNPAID)

    ensure_group_monthly_charges(start_date=start_date, end_date=end_date)
    if not can_view_all_calendar(request.user):
        prof = Professional.objects.filter(user=request.user).first()
        if prof:
            debt_qs = debt_qs.filter(professional=prof)
            group_revenue_qs = group_revenue_qs.filter(professional=prof)
            group_debt_qs = group_debt_qs.filter(professional=prof)
    if professional_id and can_view_all_calendar(request.user):
        revenue_qs = revenue_qs.filter(professional_id=professional_id)
        debt_qs = debt_qs.filter(professional_id=professional_id)
        group_revenue_qs = group_revenue_qs.filter(professional_id=professional_id)
        group_debt_qs = group_debt_qs.filter(professional_id=professional_id)
    if service_id:
        revenue_qs = revenue_qs.filter(service_id=service_id)
        debt_qs = debt_qs.filter(service_id=service_id)
        group_revenue_qs = group_revenue_qs.filter(service_id=service_id)
        group_debt_qs = group_debt_qs.filter(service_id=service_id)

    revenue_qs = revenue_qs.filter(date__gte=start_date, date__lte=end_date)
    start_month = start_date.replace(day=1)
    end_month = end_date.replace(day=1)
    group_revenue_qs = group_revenue_qs.filter(month__gte=start_month, month__lte=end_month)
    group_debt_qs = group_debt_qs.filter(month__lte=today.replace(day=1))

    appointment_total_period = (
        revenue_qs.aggregate(total=Coalesce(Sum("final_price"), Decimal("0.00")))
        .get("total")
        or Decimal("0.00")
    )
    group_total_period = (
        group_revenue_qs.aggregate(total=Coalesce(Sum("final_price"), Decimal("0.00")))
        .get("total")
        or Decimal("0.00")
    )
    total_period = appointment_total_period + group_total_period
    total_count = revenue_qs.count() + group_revenue_qs.count()
    days_span = (end_date - start_date).days + 1
    avg_per_day = total_period / Decimal(days_span) if days_span > 0 else Decimal("0.00")

    debt_totals_appts = debt_qs.aggregate(
        debt_total=Coalesce(Sum("final_price"), Decimal("0.00")),
        debt_appointments=Count("id"),
        debt_clients=Count("client_id", distinct=True),
    )
    debt_totals_groups = group_debt_qs.aggregate(
        debt_total=Coalesce(Sum("final_price"), Decimal("0.00")),
        debt_appointments=Count("id"),
        debt_clients=Count("client_id", distinct=True),
    )
    debt_total = (debt_totals_appts["debt_total"] or Decimal("0.00")) + (debt_totals_groups["debt_total"] or Decimal("0.00"))
    debt_appointments = (debt_totals_appts["debt_appointments"] or 0) + (debt_totals_groups["debt_appointments"] or 0)
    debt_clients = debt_qs.values_list("client_id", flat=True).distinct()
    debt_clients_group = group_debt_qs.values_list("client_id", flat=True).distinct()
    debt_clients_count = len(set(list(debt_clients) + list(debt_clients_group)))

    appointment_daily_rows = (
        revenue_qs.values("date")
        .annotate(total=Coalesce(Sum("final_price"), Decimal("0.00")), count=Count("id"))
        .order_by("-date")
    )
    group_daily_rows = (
        group_revenue_qs.values("month")
        .annotate(total=Coalesce(Sum("final_price"), Decimal("0.00")), count=Count("id"))
        .order_by("-month")
    )

    daily_map = {}
    for row in appointment_daily_rows:
        key = row["date"]
        existing = daily_map.get(key, {"date": key, "total": Decimal("0.00"), "count": 0})
        existing["total"] += row["total"] or Decimal("0.00")
        existing["count"] += row["count"] or 0
        daily_map[key] = existing
    for row in group_daily_rows:
        key = row["month"]
        existing = daily_map.get(key, {"date": key, "total": Decimal("0.00"), "count": 0})
        existing["total"] += row["total"] or Decimal("0.00")
        existing["count"] += row["count"] or 0
        daily_map[key] = existing
    daily_rows = sorted(daily_map.values(), key=lambda item: item["date"], reverse=True)

    if request.GET.get("export") == "1":
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = "attachment; filename=faturacao.csv"
        writer = csv.writer(response)
        writer.writerow(["Data", "Total", "Marcações"])
        for row in daily_rows:
            writer.writerow([row["date"].strftime("%Y-%m-%d"), row["total"], row["count"]])
        return response

    paginator = Paginator(daily_rows, per_page)
    page_obj = paginator.get_page(request.GET.get("page") or 1)

    return render(
        request,
        "backoffice/faturacao.html",
        {
            "page_obj": page_obj,
            "paginator": paginator,
            "per_page": per_page,
            "start": start_date.strftime("%Y-%m-%d"),
            "end": end_date.strftime("%Y-%m-%d"),
            "professional_id": professional_id,
            "service_id": service_id,
            "total_period": total_period,
            "avg_per_day": avg_per_day,
            "total_count": total_count,
            "debt_total": debt_total,
            "debt_appointments": debt_appointments,
            "debt_clients": debt_clients_count,
            "professionals": Professional.objects.select_related("user").order_by("user__username"),
            "services": Service.objects.order_by("name"),
            "return_to": request.get_full_path(),
        },
    )


def _parse_date_param(value):
    value = (value or "").strip()
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return None


@backoffice_required
def backoffice_cash_dashboard_view(request):
    if not _can_access_cash_area(request.user):
        return HttpResponseForbidden("Acesso reservado a administradores e receção.")

    session_status = (request.GET.get("session_status") or "all").strip().lower()
    session_date_value = (request.GET.get("session_date") or "").strip()
    session_date = _parse_date_param(session_date_value)

    open_session = (
        CashSession.objects.select_related("opened_by", "closed_by")
        .filter(status=CashSession.STATUS_OPEN)
        .order_by("-opened_at", "-id")
        .first()
    )
    all_sessions_qs = CashSession.objects.select_related("opened_by", "closed_by").all()
    sessions_qs = all_sessions_qs
    if session_status in {CashSession.STATUS_OPEN, CashSession.STATUS_CLOSED}:
        sessions_qs = sessions_qs.filter(status=session_status)
    if session_date:
        sessions_qs = sessions_qs.filter(session_date=session_date)
    session_id = (request.GET.get("session_id") or request.POST.get("session_id") or "").strip()
    selected_session = None
    if session_id:
        selected_session = get_object_or_404(all_sessions_qs, id=session_id)
    else:
        selected_session = open_session if open_session and (session_status in {"all", CashSession.STATUS_OPEN}) and (not session_date or open_session.session_date == session_date) else sessions_qs.first()
    active_tab = (request.GET.get("tab") or request.POST.get("tab") or "overview").strip().lower()
    if active_tab == "receipts":
        active_tab = "overview"
    if active_tab not in {"overview", "history", "closings"}:
        active_tab = "overview"

    selected_totals = _cash_session_totals(selected_session) if selected_session else {
        "total_in": Decimal("0.00"),
        "total_out": Decimal("0.00"),
        "balance": Decimal("0.00"),
        "cash_in": Decimal("0.00"),
        "cash_out": Decimal("0.00"),
        "expected_cash": Decimal("0.00"),
        "movement_count": 0,
        "voided_count": 0,
        "payment_breakdown": [],
    }
    source_breakdown = selected_totals["source_breakdown"] if selected_session else []
    pending_appointments_qs = _pending_cash_appointments_for_session(selected_session) if selected_session else Appointment.objects.none()
    pending_client_payments_qs = _pending_cash_client_payments_for_session(selected_session) if selected_session else ClientPayment.objects.none()
    pending_group_monthly_qs = _pending_cash_group_monthly_for_session(selected_session) if selected_session else GroupMonthlyCharge.objects.none()
    void_status = (request.GET.get("void_status") or "active").strip().lower()
    if void_status not in {"active", "voided", "all"}:
        void_status = "active"

    open_form = CashSessionOpenForm(initial={
        "session_date": timezone.localdate().strftime("%Y-%m-%d"),
        "opening_amount": Decimal("0.00"),
    })
    close_form = CashSessionCloseForm(initial={
        "counted_cash_amount": selected_totals["expected_cash"],
    })
    manual_form = CashManualMovementForm(initial={
        "happened_at": timezone.localtime().strftime("%Y-%m-%dT%H:%M"),
    })
    appointment_form = CashAppointmentMovementForm(appointment_queryset=pending_appointments_qs)
    client_payment_form = CashClientPaymentMovementForm(payment_queryset=pending_client_payments_qs)
    group_monthly_form = CashGroupMonthlyMovementForm(monthly_charge_queryset=pending_group_monthly_qs)
    void_form = CashVoidMovementForm()
    edit_movement_id = (request.GET.get("edit_movement_id") or "").strip()
    void_movement_id = (request.GET.get("void_movement_id") or "").strip()
    editing_movement = None
    voiding_movement = None
    if edit_movement_id and selected_session:
        editing_movement = get_object_or_404(
            selected_session.movements.filter(
                id=edit_movement_id,
                source_type=CashMovement.SOURCE_MANUAL,
                is_void=False,
            )
        )
        manual_form = CashManualMovementForm(initial={
            "movement_type": editing_movement.movement_type,
            "payment_method": editing_movement.payment_method,
            "amount": editing_movement.amount,
            "description": editing_movement.description,
            "happened_at": timezone.localtime(editing_movement.happened_at).strftime("%Y-%m-%dT%H:%M") if editing_movement.happened_at else "",
            "notes": editing_movement.notes,
        })
    if void_movement_id and selected_session:
        voiding_movement = get_object_or_404(
            selected_session.movements.filter(
                id=void_movement_id,
                source_type=CashMovement.SOURCE_MANUAL,
                is_void=False,
            )
        )

    current_base_qs = _cash_dashboard_querystring(
        request,
        selected_session,
        session_status,
        session_date_value,
        extra={"tab": active_tab},
        remove={"edit_movement_id", "void_movement_id"},
    )

    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()

        if action == "open_session":
            open_form = CashSessionOpenForm(request.POST)
            if open_session:
                messages.error(request, "Já existe uma sessão de caixa aberta.")
            elif open_form.is_valid():
                session = CashSession.objects.create(
                    session_date=open_form.cleaned_data["session_date"],
                    opening_amount=open_form.cleaned_data["opening_amount"],
                    expected_cash_amount=open_form.cleaned_data["opening_amount"],
                    opening_notes=open_form.cleaned_data.get("opening_notes") or "",
                    opened_by=request.user,
                )
                log_audit_event(
                    category="cash",
                    action="session_opened",
                    request=request,
                    actor=request.user,
                    instance=session,
                    source="backoffice_cash",
                    message="Sessão de caixa aberta.",
                    after=snapshot_instance(
                        session,
                        fields=["session_date", "status", "opening_amount", "expected_cash_amount", "opened_by_id", "opened_at"],
                    ),
                )
                messages.success(request, "Sessão de caixa aberta com sucesso.")
                return redirect(f"{reverse('backoffice_cash_dashboard')}?session_id={session.id}")

        elif action == "close_session":
            if not selected_session or selected_session.status != CashSession.STATUS_OPEN:
                messages.error(request, "Seleciona uma sessão de caixa aberta.")
            else:
                close_form = CashSessionCloseForm(request.POST)
                if close_form.is_valid():
                    before = snapshot_instance(
                        selected_session,
                        fields=[
                            "status",
                            "expected_cash_amount",
                            "counted_cash_amount",
                            "difference_amount",
                            "closed_by_id",
                            "closed_at",
                        ],
                    )
                    expected_cash = selected_totals["expected_cash"]
                    counted_cash = close_form.cleaned_data["counted_cash_amount"]
                    difference = counted_cash - expected_cash
                    selected_session.expected_cash_amount = expected_cash
                    selected_session.counted_cash_amount = counted_cash
                    selected_session.difference_amount = difference
                    selected_session.closing_notes = close_form.cleaned_data.get("closing_notes") or ""
                    selected_session.closed_by = request.user
                    selected_session.closed_at = timezone.now()
                    selected_session.status = CashSession.STATUS_CLOSED
                    selected_session.save(
                        update_fields=[
                            "expected_cash_amount",
                            "counted_cash_amount",
                            "difference_amount",
                            "closing_notes",
                            "closed_by",
                            "closed_at",
                            "status",
                        ]
                    )
                    log_audit_event(
                        category="cash",
                        action="session_closed",
                        request=request,
                        actor=request.user,
                        instance=selected_session,
                        source="backoffice_cash",
                        message="Sessão de caixa fechada.",
                        before=before,
                        after=snapshot_instance(
                            selected_session,
                            fields=[
                                "status",
                                "expected_cash_amount",
                                "counted_cash_amount",
                                "difference_amount",
                                "closed_by_id",
                                "closed_at",
                            ],
                        ),
                    )
                    messages.success(request, "Sessão de caixa fechada com sucesso.")
                    return redirect(f"{reverse('backoffice_cash_dashboard')}?session_id={selected_session.id}")

        elif action == "reopen_session":
            if not is_admin_role(request.user):
                messages.error(request, "Só administradores podem reabrir sessões de caixa.")
            elif not selected_session or selected_session.status != CashSession.STATUS_CLOSED:
                messages.error(request, "Seleciona uma sessão de caixa fechada.")
            elif open_session and open_session.id != selected_session.id:
                messages.error(request, "Já existe outra sessão de caixa aberta.")
            else:
                before = snapshot_instance(
                    selected_session,
                    fields=[
                        "status",
                        "expected_cash_amount",
                        "counted_cash_amount",
                        "difference_amount",
                        "closing_notes",
                        "closed_by_id",
                        "closed_at",
                    ],
                )
                selected_session.status = CashSession.STATUS_OPEN
                selected_session.counted_cash_amount = None
                selected_session.difference_amount = Decimal("0.00")
                selected_session.closing_notes = ""
                selected_session.closed_by = None
                selected_session.closed_at = None
                selected_session.save(
                    update_fields=[
                        "status",
                        "counted_cash_amount",
                        "difference_amount",
                        "closing_notes",
                        "closed_by",
                        "closed_at",
                    ]
                )
                log_audit_event(
                    category="cash",
                    action="session_reopened",
                    request=request,
                    actor=request.user,
                    instance=selected_session,
                    source="backoffice_cash",
                    message="Sessão de caixa reaberta.",
                    before=before,
                    after=snapshot_instance(
                        selected_session,
                        fields=[
                            "status",
                            "expected_cash_amount",
                            "counted_cash_amount",
                            "difference_amount",
                            "closing_notes",
                            "closed_by_id",
                            "closed_at",
                        ],
                    ),
                )
                messages.success(request, "Sessão de caixa reaberta com sucesso.")
                return redirect(f"{reverse('backoffice_cash_dashboard')}?session_id={selected_session.id}")

        elif action == "add_manual_movement":
            if not selected_session or selected_session.status != CashSession.STATUS_OPEN:
                messages.error(request, "Seleciona uma sessão de caixa aberta.")
            else:
                manual_form = CashManualMovementForm(request.POST)
                if manual_form.is_valid():
                    movement = CashMovement.objects.create(
                        session=selected_session,
                        movement_type=manual_form.cleaned_data["movement_type"],
                        source_type=CashMovement.SOURCE_MANUAL,
                        payment_method=manual_form.cleaned_data["payment_method"],
                        amount=manual_form.cleaned_data["amount"],
                        description=manual_form.cleaned_data["description"],
                        notes=manual_form.cleaned_data.get("notes") or "",
                        created_by=request.user,
                        happened_at=manual_form.cleaned_data["happened_at"],
                    )
                    log_audit_event(
                        category="cash_movement",
                        action="manual_created",
                        request=request,
                        actor=request.user,
                        instance=movement,
                        source="backoffice_cash",
                        message="Movimento manual de caixa registado.",
                        after={
                            "session_id": selected_session.id,
                            "movement_type": movement.movement_type,
                            "payment_method": movement.payment_method,
                            "amount": str(movement.amount),
                            "description": movement.description,
                            "happened_at": movement.happened_at.isoformat() if movement.happened_at else "",
                        },
                    )
                    messages.success(request, "Movimento manual registado com sucesso.")
                    return redirect(f"{reverse('backoffice_cash_dashboard')}?{current_base_qs}")

        elif action == "update_manual_movement":
            movement_id = (request.POST.get("movement_id") or "").strip()
            if not selected_session or selected_session.status != CashSession.STATUS_OPEN:
                messages.error(request, "Seleciona uma sessão de caixa aberta.")
            else:
                editing_movement = get_object_or_404(
                    selected_session.movements.filter(
                        id=movement_id,
                        source_type=CashMovement.SOURCE_MANUAL,
                        is_void=False,
                    )
                )
                manual_form = CashManualMovementForm(request.POST)
                if manual_form.is_valid():
                    before = {
                        "session_id": selected_session.id,
                        "movement_type": editing_movement.movement_type,
                        "payment_method": editing_movement.payment_method,
                        "amount": str(editing_movement.amount),
                        "description": editing_movement.description,
                        "notes": editing_movement.notes,
                        "happened_at": editing_movement.happened_at.isoformat() if editing_movement.happened_at else "",
                    }
                    editing_movement.movement_type = manual_form.cleaned_data["movement_type"]
                    editing_movement.payment_method = manual_form.cleaned_data["payment_method"]
                    editing_movement.amount = manual_form.cleaned_data["amount"]
                    editing_movement.description = manual_form.cleaned_data["description"]
                    editing_movement.notes = manual_form.cleaned_data.get("notes") or ""
                    editing_movement.happened_at = manual_form.cleaned_data["happened_at"]
                    editing_movement.full_clean()
                    editing_movement.save(
                        update_fields=[
                            "movement_type",
                            "payment_method",
                            "amount",
                            "description",
                            "notes",
                            "happened_at",
                        ]
                    )
                    log_audit_event(
                        category="cash_movement",
                        action="manual_updated",
                        request=request,
                        actor=request.user,
                        instance=editing_movement,
                        source="backoffice_cash",
                        message="Movimento manual de caixa atualizado.",
                        before=before,
                        after={
                            "session_id": selected_session.id,
                            "movement_type": editing_movement.movement_type,
                            "payment_method": editing_movement.payment_method,
                            "amount": str(editing_movement.amount),
                            "description": editing_movement.description,
                            "notes": editing_movement.notes,
                            "happened_at": editing_movement.happened_at.isoformat() if editing_movement.happened_at else "",
                        },
                    )
                    messages.success(request, "Movimento manual atualizado com sucesso.")
                    return redirect(f"{reverse('backoffice_cash_dashboard')}?{current_base_qs}")

        elif action == "void_manual_movement":
            movement_id = (request.POST.get("movement_id") or "").strip()
            if not selected_session or selected_session.status != CashSession.STATUS_OPEN:
                messages.error(request, "Seleciona uma sessão de caixa aberta.")
            else:
                movement = get_object_or_404(
                    selected_session.movements.filter(
                        id=movement_id,
                        source_type=CashMovement.SOURCE_MANUAL,
                        is_void=False,
                    )
                )
                voiding_movement = movement
                void_form = CashVoidMovementForm(request.POST)
                if void_form.is_valid():
                    before = {
                        "session_id": selected_session.id,
                        "movement_type": movement.movement_type,
                        "payment_method": movement.payment_method,
                        "amount": str(movement.amount),
                        "description": movement.description,
                        "notes": movement.notes,
                        "happened_at": movement.happened_at.isoformat() if movement.happened_at else "",
                    }
                    movement.is_void = True
                    movement.void_reason = void_form.cleaned_data["void_reason"]
                    movement.voided_by = request.user
                    movement.voided_at = timezone.now()
                    movement.save(update_fields=["is_void", "void_reason", "voided_by", "voided_at"])
                    log_audit_event(
                        category="cash_movement",
                        action="manual_voided",
                        request=request,
                        actor=request.user,
                        instance=movement,
                        source="backoffice_cash",
                        message="Movimento manual de caixa anulado.",
                        before=before,
                        after={
                            "session_id": selected_session.id,
                            "is_void": True,
                            "void_reason": movement.void_reason,
                            "voided_by_id": movement.voided_by_id,
                            "voided_at": movement.voided_at.isoformat() if movement.voided_at else "",
                        },
                    )
                    messages.success(request, "Movimento manual anulado com sucesso.")
                    return redirect(f"{reverse('backoffice_cash_dashboard')}?{current_base_qs}")

        elif action == "add_paid_appointment":
            if not selected_session or selected_session.status != CashSession.STATUS_OPEN:
                messages.error(request, "Seleciona uma sessão de caixa aberta.")
            else:
                appointment_form = CashAppointmentMovementForm(
                    request.POST,
                    appointment_queryset=pending_appointments_qs,
                )
                if appointment_form.is_valid():
                    appointment = appointment_form.cleaned_data["appointment"]
                    client_name = (
                        getattr(getattr(appointment.client, "client_profile", None), "full_name", "")
                        or appointment.client.get_full_name()
                        or appointment.client.username
                    )
                    movement = CashMovement.objects.create(
                        session=selected_session,
                        movement_type=CashMovement.TYPE_IN,
                        source_type=CashMovement.SOURCE_APPOINTMENT,
                        payment_method=appointment_form.cleaned_data["payment_method"],
                        amount=appointment.final_price,
                        description=f"Marcação · {client_name} · {appointment.service.name}",
                        notes=appointment_form.cleaned_data.get("notes") or "",
                        appointment=appointment,
                        created_by=request.user,
                        happened_at=appointment.paid_at or timezone.now(),
                    )
                    log_audit_event(
                        category="cash_movement",
                        action="appointment_created",
                        request=request,
                        actor=request.user,
                        instance=movement,
                        source="backoffice_cash",
                        message="Recebimento de marcação lançado em caixa.",
                        after={
                            "session_id": selected_session.id,
                            "appointment_id": appointment.id,
                            "payment_method": movement.payment_method,
                            "amount": str(movement.amount),
                        },
                    )
                    messages.success(request, "Recebimento de marcação lançado com sucesso.")
                    return redirect(f"{reverse('backoffice_cash_dashboard')}?{current_base_qs}")

        elif action == "add_client_payment":
            if not selected_session or selected_session.status != CashSession.STATUS_OPEN:
                messages.error(request, "Seleciona uma sessão de caixa aberta.")
            else:
                client_payment_form = CashClientPaymentMovementForm(
                    request.POST,
                    payment_queryset=pending_client_payments_qs,
                )
                if client_payment_form.is_valid():
                    payment = client_payment_form.cleaned_data["client_payment"]
                    movement, movement_created, movement_error = ensure_client_payment_cash_movement(
                        payment,
                        session=selected_session,
                        notes_append=client_payment_form.cleaned_data.get("notes") or "",
                    )
                    if movement_created and movement:
                        log_audit_event(
                            category="cash_movement",
                            action="client_payment_created",
                            request=request,
                            actor=request.user,
                            instance=movement,
                            source="backoffice_cash",
                            message="Pagamento de cliente lançado em caixa.",
                            after={
                                "session_id": selected_session.id,
                                "client_payment_id": payment.id,
                                "payment_method": movement.payment_method,
                                "amount": str(movement.amount),
                            },
                        )
                        messages.success(request, "Pagamento de cliente lançado em caixa com sucesso.")
                        return redirect(f"{reverse('backoffice_cash_dashboard')}?{current_base_qs}")
                    if movement and not movement_created:
                        messages.info(request, "Este pagamento já estava lançado em caixa.")
                        return redirect(f"{reverse('backoffice_cash_dashboard')}?{current_base_qs}")
                    messages.error(request, movement_error or "Não foi possível lançar o pagamento em caixa.")

        elif action == "add_paid_group_monthly":
            if not selected_session or selected_session.status != CashSession.STATUS_OPEN:
                messages.error(request, "Seleciona uma sessão de caixa aberta.")
            else:
                group_monthly_form = CashGroupMonthlyMovementForm(
                    request.POST,
                    monthly_charge_queryset=pending_group_monthly_qs,
                )
                if group_monthly_form.is_valid():
                    charge = group_monthly_form.cleaned_data["group_monthly_charge"]
                    client_name = (
                        getattr(getattr(charge.client, "client_profile", None), "full_name", "")
                        or charge.client.get_full_name()
                        or charge.client.username
                    )
                    class_name = charge.class_name or getattr(charge.service, "name", "") or "Turma"
                    movement = CashMovement.objects.create(
                        session=selected_session,
                        movement_type=CashMovement.TYPE_IN,
                        source_type=CashMovement.SOURCE_GROUP_MONTHLY,
                        payment_method=group_monthly_form.cleaned_data["payment_method"],
                        amount=charge.final_price,
                        description=f"Turma · {client_name} · {class_name}",
                        notes=group_monthly_form.cleaned_data.get("notes") or "",
                        group_monthly_charge=charge,
                        created_by=request.user,
                        happened_at=charge.paid_at or timezone.now(),
                    )
                    log_audit_event(
                        category="cash_movement",
                        action="group_monthly_created",
                        request=request,
                        actor=request.user,
                        instance=movement,
                        source="backoffice_cash",
                        message="Recebimento de mensalidade de turma lançado em caixa.",
                        after={
                            "session_id": selected_session.id,
                            "group_monthly_charge_id": charge.id,
                            "payment_method": movement.payment_method,
                            "amount": str(movement.amount),
                        },
                    )
                    messages.success(request, "Recebimento de turma lançado com sucesso.")
                    return redirect(f"{reverse('backoffice_cash_dashboard')}?{current_base_qs}")

        else:
            messages.error(request, "Ação inválida.")

    movement_type = (request.GET.get("movement_type") or "all").strip()
    payment_method = (request.GET.get("payment_method") or "all").strip()
    source_type = (request.GET.get("source_type") or "all").strip()
    movement_start = (request.GET.get("movement_start") or "").strip()
    movement_end = (request.GET.get("movement_end") or "").strip()
    movement_start_dt = None
    movement_end_dt = None
    try:
        movement_per_page = int(request.GET.get("movement_per_page") or 15)
    except (TypeError, ValueError):
        movement_per_page = 15
    if movement_per_page not in (15, 30, 50, 100):
        movement_per_page = 15

    if movement_start:
        try:
            movement_start_dt = datetime.strptime(movement_start, "%Y-%m-%d").date()
        except ValueError:
            movement_start_dt = None
    if movement_end:
        try:
            movement_end_dt = datetime.strptime(movement_end, "%Y-%m-%d").date()
        except ValueError:
            movement_end_dt = None

    movement_qs = (
        selected_session.movements
        .select_related(
            "created_by",
            "appointment",
            "appointment__client",
            "appointment__client__client_profile",
            "appointment__service",
            "client_profile",
            "group_monthly_charge",
            "group_monthly_charge__client",
            "group_monthly_charge__client__client_profile",
            "group_monthly_charge__service",
            "stock_movement",
            "voided_by",
        )
        if selected_session
        else CashMovement.objects.none()
    )
    if void_status == "active":
        movement_qs = movement_qs.filter(is_void=False)
    elif void_status == "voided":
        movement_qs = movement_qs.filter(is_void=True)
    if movement_type in {CashMovement.TYPE_IN, CashMovement.TYPE_OUT}:
        movement_qs = movement_qs.filter(movement_type=movement_type)
    if payment_method in dict(CashMovement.PAYMENT_METHOD_CHOICES):
        movement_qs = movement_qs.filter(payment_method=payment_method)
    if source_type in dict(CashMovement.SOURCE_CHOICES):
        movement_qs = movement_qs.filter(source_type=source_type)
    if movement_start_dt:
        movement_qs = movement_qs.filter(happened_at__date__gte=movement_start_dt)
    if movement_end_dt:
        movement_qs = movement_qs.filter(happened_at__date__lte=movement_end_dt)

    if (request.GET.get("export") or "").strip().lower() == "csv" and selected_session:
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        filename = f"caixa-{selected_session.session_date.strftime('%Y%m%d')}-sessao-{selected_session.id}.csv"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        writer = csv.writer(response)
        writer.writerow([
            "Sessão",
            "Quando",
            "Tipo",
            "Origem",
            "Método",
            "Descrição",
            "Valor",
            "Estado",
            "Motivo anulação",
            "Utente",
            "Marcação",
            "Mensalidade",
            "Stock",
            "Criado por",
        ])
        for movement in movement_qs.order_by("happened_at", "id"):
            writer.writerow([
                selected_session.session_date.strftime("%Y-%m-%d"),
                timezone.localtime(movement.happened_at).strftime("%Y-%m-%d %H:%M") if movement.happened_at else "",
                movement.get_movement_type_display(),
                movement.get_source_type_display(),
                movement.get_payment_method_display(),
                movement.description,
                movement.amount,
                "Anulado" if movement.is_void else "Ativo",
                movement.void_reason,
                movement.client_profile.full_name if movement.client_profile_id else "",
                movement.appointment_id or "",
                movement.group_monthly_charge_id or "",
                movement.stock_movement_id or "",
                movement.created_by.get_full_name() if movement.created_by else "",
            ])
        return response

    if (request.GET.get("export") or "").strip().lower() == "closing_csv" and selected_session:
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        filename = f"caixa-fecho-{selected_session.session_date.strftime('%Y%m%d')}-sessao-{selected_session.id}.csv"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        writer = csv.writer(response)
        writer.writerow(["Campo", "Valor"])
        writer.writerow(["Sessão", selected_session.id])
        writer.writerow(["Data", selected_session.session_date.strftime("%Y-%m-%d")])
        writer.writerow(["Estado", selected_session.get_status_display()])
        writer.writerow(["Abertura", selected_session.opening_amount])
        writer.writerow(["Entradas", selected_totals["total_in"]])
        writer.writerow(["Saídas", selected_totals["total_out"]])
        writer.writerow(["Saldo teórico", selected_totals["balance"]])
        writer.writerow(["Numerário esperado", selected_totals["expected_cash"]])
        writer.writerow(["Numerário contado", selected_session.counted_cash_amount or Decimal("0.00")])
        writer.writerow(["Diferença", selected_session.difference_amount])
        writer.writerow(["Movimentos ativos", selected_totals["movement_count"]])
        writer.writerow(["Movimentos anulados", selected_totals["voided_count"]])
        writer.writerow([])
        writer.writerow(["Origem", "Total", "Movimentos"])
        for row in selected_totals["source_breakdown"]:
            writer.writerow([row["label"], row["total"], row["count"]])
        writer.writerow([])
        writer.writerow(["Método", "Saldo"])
        for row in selected_totals["payment_breakdown"]:
            writer.writerow([row["label"], row["net"]])
        return response

    current_view_params = request.GET.copy()
    current_view_params.pop("export", None)
    current_view_params.pop("edit_movement_id", None)
    current_view_params.pop("void_movement_id", None)
    current_view_qs = current_view_params.urlencode()
    tab_base_params = current_view_params.copy()
    tab_base_params.pop("tab", None)
    tab_base_qs = tab_base_params.urlencode()

    movement_filter_params = request.GET.copy()
    movement_filter_params.pop("movement_page", None)
    movement_filter_params.pop("export", None)
    movement_filter_params.pop("edit_movement_id", None)
    movement_filter_params.pop("void_movement_id", None)
    movement_filter_qs = movement_filter_params.urlencode()
    export_qs = movement_filter_params.copy()
    export_qs["export"] = "csv"
    closing_export_qs = movement_filter_params.copy()
    closing_export_qs["export"] = "closing_csv"

    movement_page_obj = Paginator(movement_qs, movement_per_page).get_page(request.GET.get("movement_page") or 1)
    recent_sessions = list(sessions_qs[:12])
    closed_sessions_qs = sessions_qs.filter(status=CashSession.STATUS_CLOSED)
    closing_page_obj = Paginator(closed_sessions_qs, 15).get_page(request.GET.get("closing_page") or 1)
    closing_filter_params = request.GET.copy()
    closing_filter_params.pop("closing_page", None)
    closing_filter_params.pop("export", None)
    closing_filter_params.pop("edit_movement_id", None)
    closing_filter_params.pop("void_movement_id", None)
    closing_filter_qs = closing_filter_params.urlencode()
    pending_appointments_preview = list(pending_appointments_qs[:8]) if selected_session else []
    for appt in pending_appointments_preview:
        appt.client_label = (
            getattr(getattr(appt.client, "client_profile", None), "full_name", "")
            or appt.client.get_full_name()
            or appt.client.username
        )
    pending_client_payments_preview = list(pending_client_payments_qs[:8]) if selected_session else []
    for payment in pending_client_payments_preview:
        payment.client_label = payment.client_profile.full_name if payment.client_profile else "Cliente"
        payment.method_label = dict(ClientPayment.PAYMENT_METHOD_CHOICES).get(payment.payment_method, payment.payment_method)
    pending_group_monthly_preview = list(pending_group_monthly_qs[:8]) if selected_session else []
    for charge in pending_group_monthly_preview:
        charge.client_label = (
            getattr(getattr(charge.client, "client_profile", None), "full_name", "")
            or charge.client.get_full_name()
            or charge.client.username
        )
        charge.class_label = charge.class_name or getattr(charge.service, "name", "") or "Turma"
    monthly_summary = _cash_month_summary(selected_session.session_date if selected_session else timezone.localdate())
    session_recent_movements = list(
        movement_qs.select_related(
            "created_by",
            "client_profile",
            "appointment",
            "appointment__client",
            "appointment__client__client_profile",
            "group_monthly_charge",
            "group_monthly_charge__client",
            "group_monthly_charge__client__client_profile",
        )[:8]
    ) if selected_session else []

    return render(
        request,
        "backoffice/cash_dashboard.html",
        {
            "selected_session": selected_session,
            "open_session": open_session,
            "selected_totals": selected_totals,
            "movement_page_obj": movement_page_obj,
            "closing_page_obj": closing_page_obj,
            "movement_filter_qs": movement_filter_qs,
            "closing_filter_qs": closing_filter_qs,
            "current_view_qs": current_view_qs,
            "tab_base_qs": tab_base_qs,
            "export_qs": export_qs.urlencode(),
            "closing_export_qs": closing_export_qs.urlencode(),
            "active_tab": active_tab,
            "movement_per_page": movement_per_page,
            "movement_type": movement_type,
            "payment_method": payment_method,
            "source_type": source_type,
            "void_status": void_status,
            "movement_start": movement_start,
            "movement_end": movement_end,
            "recent_sessions": recent_sessions,
            "pending_appointments_count": pending_appointments_qs.count() if selected_session else 0,
            "pending_appointments_preview": pending_appointments_preview,
            "pending_client_payments_count": pending_client_payments_qs.count() if selected_session else 0,
            "pending_client_payments_preview": pending_client_payments_preview,
            "pending_group_monthly_count": pending_group_monthly_qs.count() if selected_session else 0,
            "pending_group_monthly_preview": pending_group_monthly_preview,
            "source_breakdown": source_breakdown,
            "monthly_summary": monthly_summary,
            "session_recent_movements": session_recent_movements,
            "open_form": open_form,
            "close_form": close_form,
            "manual_form": manual_form,
            "appointment_form": appointment_form,
            "client_payment_form": client_payment_form,
            "group_monthly_form": group_monthly_form,
            "void_form": void_form,
            "editing_movement": editing_movement,
            "voiding_movement": voiding_movement,
            "session_status": session_status,
            "session_date": session_date_value,
            "cash_session_status_choices": CashSession.STATUS_CHOICES,
            "cash_movement_type_choices": CashMovement.TYPE_CHOICES,
            "cash_payment_method_choices": CashMovement.PAYMENT_METHOD_CHOICES,
            "cash_source_type_choices": CashMovement.SOURCE_CHOICES,
            "can_reopen_session": is_admin_role(request.user),
        },
    )


@backoffice_required
def backoffice_cash_session_print_view(request, session_id):
    if not _can_access_cash_area(request.user):
        return HttpResponseForbidden("Acesso reservado a administradores e receção.")

    session = get_object_or_404(
        CashSession.objects.select_related("opened_by", "closed_by"),
        id=session_id,
    )
    totals = _cash_session_totals(session)
    movements = list(
        session.movements
        .select_related(
            "created_by",
            "voided_by",
            "appointment",
            "appointment__client",
            "appointment__client__client_profile",
            "appointment__service",
            "client_profile",
            "group_monthly_charge",
            "group_monthly_charge__client",
            "group_monthly_charge__client__client_profile",
            "group_monthly_charge__service",
            "stock_movement",
        )
        .order_by("happened_at", "id")
    )
    return render(
        request,
        "backoffice/cash_session_print.html",
        {
            "selected_session": session,
            "selected_totals": totals,
            "movements": movements,
            "generated_at": timezone.localtime(),
        },
    )


def _subcontractor_lines_queryset(request):
    qs = (
        SubcontractorPaymentLine.objects
        .select_related("professional", "professional__user", "client", "service", "appointment")
    )
    today = timezone.localdate()
    period = (request.GET.get("period") or "month").strip().lower()
    if period not in {"month", "total", "custom"}:
        period = "month"
    start_param = (request.GET.get("start") or "").strip()
    end_param = (request.GET.get("end") or "").strip()
    professional_id = (request.GET.get("professional_id") or "").strip()
    service_id = (request.GET.get("service_id") or "").strip()
    status = (request.GET.get("status") or "").strip()

    start_date = None
    end_date = None
    start_value = ""
    end_value = ""
    if period == "month":
        start_date = today - timedelta(days=29)
        end_date = today
        start_value = start_date.strftime("%Y-%m-%d")
        end_value = end_date.strftime("%Y-%m-%d")
        qs = qs.filter(appointment_date__gte=start_date, appointment_date__lte=end_date)
    elif period == "custom":
        start_date = _parse_date_param(start_param) or (today - timedelta(days=29))
        end_date = _parse_date_param(end_param) or today
        if start_date > end_date:
            start_date, end_date = end_date, start_date
        start_value = start_date.strftime("%Y-%m-%d")
        end_value = end_date.strftime("%Y-%m-%d")
        qs = qs.filter(appointment_date__gte=start_date, appointment_date__lte=end_date)

    if professional_id:
        qs = qs.filter(professional_id=professional_id)
    if service_id:
        qs = qs.filter(service_id=service_id)
    if status in {
        SubcontractorPaymentLine.STATUS_UNPAID,
        SubcontractorPaymentLine.STATUS_PAID,
        SubcontractorPaymentLine.STATUS_VOID,
    }:
        qs = qs.filter(status=status)

    return qs, {
        "period": period,
        "start": start_value,
        "end": end_value,
        "start_date": start_date,
        "end_date": end_date,
        "professional_id": professional_id,
        "service_id": service_id,
        "status": status,
    }


def backoffice_subcontractors_view(request):
    if not is_admin_role(request.user):
        return HttpResponseForbidden("Acesso reservado a administradores.")

    qs, filters = _subcontractor_lines_queryset(request)

    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()
        ids = request.POST.getlist("line_ids")
        if not ids:
            messages.error(request, "Seleciona pelo menos uma linha.")
            return redirect(request.get_full_path())
        selected = list(
            SubcontractorPaymentLine.objects.select_related("professional", "appointment").filter(id__in=ids)
        )
        if action == "mark_paid":
            now = timezone.now()
            SubcontractorPaymentLine.objects.filter(id__in=[line.id for line in selected]).update(
                status=SubcontractorPaymentLine.STATUS_PAID,
                paid_at=now,
                paid_by=request.user,
            )
            for line in selected:
                log_audit_event(
                    category="subcontractors",
                    action="mark_paid",
                    request=request,
                    actor=request.user,
                    instance=line,
                    source="backoffice_subcontractors",
                    message="Linha de subcontratado marcada como paga.",
                    before={
                        "status": line.status,
                        "paid_at": line.paid_at,
                        "paid_by_id": line.paid_by_id,
                    },
                    after={
                        "status": SubcontractorPaymentLine.STATUS_PAID,
                        "paid_at": now,
                        "paid_by_id": request.user.id,
                    },
                )
            messages.success(request, "Linhas marcadas como pagas.")
        elif action == "mark_unpaid":
            SubcontractorPaymentLine.objects.filter(id__in=[line.id for line in selected]).update(
                status=SubcontractorPaymentLine.STATUS_UNPAID,
                paid_at=None,
                paid_by=None,
            )
            for line in selected:
                log_audit_event(
                    category="subcontractors",
                    action="mark_unpaid",
                    request=request,
                    actor=request.user,
                    instance=line,
                    source="backoffice_subcontractors",
                    message="Linha de subcontratado marcada como em aberto.",
                    before={
                        "status": line.status,
                        "paid_at": line.paid_at,
                        "paid_by_id": line.paid_by_id,
                    },
                    after={
                        "status": SubcontractorPaymentLine.STATUS_UNPAID,
                        "paid_at": None,
                        "paid_by_id": None,
                    },
                )
            messages.success(request, "Linhas marcadas como em aberto.")
        else:
            messages.error(request, "Ação inválida.")
        return redirect(request.get_full_path())

    per_page = request.GET.get("per_page") or "10"
    try:
        per_page = int(per_page)
    except (TypeError, ValueError):
        per_page = 10
    per_page_options = [5, 10, 15, 25, 50]
    if per_page not in per_page_options:
        per_page = 10

    totals_base = qs.exclude(status=SubcontractorPaymentLine.STATUS_VOID)
    total_period = (
        totals_base
        .aggregate(total=Coalesce(Sum("payable_amount"), Decimal("0.00")))
        .get("total")
        or Decimal("0.00")
    )
    total_paid = (
        totals_base.filter(status=SubcontractorPaymentLine.STATUS_PAID)
        .aggregate(total=Coalesce(Sum("payable_amount"), Decimal("0.00")))
        .get("total")
        or Decimal("0.00")
    )
    total_unpaid = (
        totals_base.filter(status=SubcontractorPaymentLine.STATUS_UNPAID)
        .aggregate(total=Coalesce(Sum("payable_amount"), Decimal("0.00")))
        .get("total")
        or Decimal("0.00")
    )
    total_count = totals_base.count()
    avg_per_day = None
    if filters["start_date"] and filters["end_date"]:
        days_span = (filters["end_date"] - filters["start_date"]).days + 1
        avg_per_day = total_period / Decimal(days_span) if days_span > 0 else Decimal("0.00")

    paginator = Paginator(qs.order_by("-appointment_date", "-appointment_time"), per_page)
    page_obj = paginator.get_page(request.GET.get("page") or 1)
    pagination_params = request.GET.copy()
    pagination_params.pop("page", None)
    export_qs = urlencode(
        {
            "start": filters["start"],
            "end": filters["end"],
            "period": filters["period"],
            "professional_id": filters["professional_id"],
            "service_id": filters["service_id"],
            "status": filters["status"],
        }
    )

    return render(
        request,
        "backoffice/subcontractors_payments_list.html",
        {
            "lines": page_obj.object_list,
            "page_obj": page_obj,
            "paginator": paginator,
            "per_page": per_page,
            "per_page_options": per_page_options,
            "filters": filters,
            "export_qs": export_qs,
            "pagination_qs": pagination_params.urlencode(),
            "total_period": total_period,
            "avg_per_day": avg_per_day,
            "total_count": total_count,
            "total_paid": total_paid,
            "total_unpaid": total_unpaid,
            "total_open": total_unpaid,
            "professionals": Professional.objects.select_related("user").order_by("user__username"),
            "services": Service.objects.order_by("name"),
            "status_choices": SubcontractorPaymentLine.STATUS_CHOICES,
            "return_to": request.get_full_path(),
        },
    )


def backoffice_subcontractors_export_view(request):
    if not is_admin_role(request.user):
        return HttpResponseForbidden("Acesso reservado a administradores.")

    qs, filters = _subcontractor_lines_queryset(request)
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = "attachment; filename=subcontratados.csv"
    writer = csv.writer(response)
    writer.writerow(
        [
            "Data",
            "Hora",
            "Utente",
            "Serviço",
            "Profissional",
            "Preço final",
            "Percentagem",
            "Valor a pagar",
            "Pago",
            "Pago em",
        ]
    )
    for line in qs.order_by("appointment_date", "appointment_time"):
        writer.writerow(
            [
                line.appointment_date.strftime("%Y-%m-%d"),
                line.appointment_time.strftime("%H:%M") if line.appointment_time else "",
                line.client.full_name if line.client else "",
                line.service.name if line.service else "",
                line.professional.user.get_full_name() if line.professional else "",
                line.gross_amount,
                line.percentage,
                line.payable_amount,
                "sim" if line.status == SubcontractorPaymentLine.STATUS_PAID else "não",
                line.paid_at.strftime("%Y-%m-%d %H:%M") if line.paid_at else "",
            ]
        )
    return response


def backoffice_clients_quick_view(request):
    return redirect("professional_clients")
    q = (request.GET.get("q") or "").strip()
    per_page = request.GET.get("per_page") or "10"
    try:
        per_page = int(per_page)
    except (TypeError, ValueError):
        per_page = 10
    if per_page not in (10, 25, 50):
        per_page = 10

    qs = ClientProfile.objects.select_related("user").order_by("full_name")
    if q:
        qs = apply_terms_filter(
            qs,
            q,
            [
                "full_name__icontains",
                "nif__icontains",
                "phone__icontains",
                "user__username__icontains",
                "user__first_name__icontains",
                "user__last_name__icontains",
                "user__email__icontains",
            ],
        )

    paginator = Paginator(qs, per_page)
    page_obj = paginator.get_page(request.GET.get("page") or 1)

    return render(
        request,
        "backoffice/clients_quick.html",
        {
            "clients": page_obj.object_list,
            "page_obj": page_obj,
            "paginator": paginator,
            "q": q,
            "per_page": per_page,
            "return_to": request.get_full_path(),
        },
    )


def backoffice_api_clients_search(request):
    if not can_book_for_any_professional(request.user):
        return HttpResponseForbidden("Acesso negado.")
    q = (request.GET.get("q") or "").strip()
    if len(q) < 2:
        return JsonResponse({"results": []})

    qs = ClientProfile.objects.select_related("user")
    qs = apply_terms_filter(
        qs,
        q,
        [
            "full_name__icontains",
            "nif__icontains",
            "phone__icontains",
            "user__username__icontains",
            "user__first_name__icontains",
            "user__last_name__icontains",
            "user__email__icontains",
        ],
    ).order_by("full_name")[:10]

    results = []
    for c in qs:
        user = c.user
        label = c.full_name or (user.get_full_name() if user else "") or (user.username if user else "—")
        results.append(
            {
                "id": c.id,
                "label": label,
                "nif": c.nif or "",
                "phone": c.phone or "",
                "has_user": bool(user),
                "user_id": user.id if user else None,
            }
        )
    return JsonResponse({"results": results})


def backoffice_api_professionals_by_service(request):
    if not can_book_for_any_professional(request.user):
        return HttpResponseForbidden("Acesso negado.")
    service_id = (request.GET.get("service_id") or "").strip()
    if not service_id:
        return JsonResponse({"results": []})
    qs = (
        Professional.objects.select_related("user")
        .filter(services__id=service_id)
        .distinct()
        .order_by("user__username")
    )
    results = [
        {"id": p.id, "label": p.user.get_full_name() or p.user.username}
        for p in qs
    ]
    return JsonResponse({"results": results})


def backoffice_api_slots(request):
    if not can_book_for_any_professional(request.user):
        return HttpResponseForbidden("Acesso negado.")
    service_id = (request.GET.get("service_id") or "").strip()
    professional_id = (request.GET.get("professional_id") or "").strip()
    date_str = (request.GET.get("date") or "").strip()

    if not (service_id and professional_id and date_str):
        return JsonResponse({"ok": False, "slots": [], "message": "Dados incompletos."})

    service = Service.objects.filter(id=service_id).first()
    prof = Professional.objects.filter(id=professional_id).first()
    if not service or not prof:
        return JsonResponse({"ok": False, "slots": [], "message": "Dados inválidos."})

    if not Professional.objects.filter(id=prof.id, services__id=service.id).exists():
        return JsonResponse({"ok": False, "slots": [], "message": "Profissional inválido para este serviço."})

    try:
        date_obj = datetime.strptime(date_str, "%Y-%m-%d").date()
    except Exception:
        return JsonResponse({"ok": False, "slots": [], "message": "Data inválida."})

    today = timezone.localdate()
    if date_obj < today:
        return JsonResponse({"ok": False, "slots": [], "message": "Não podes marcar no passado."})

    if not professional_works_on_date(prof, date_obj):
        days = professional_weekdays_labels(prof)
        return JsonResponse(
            {
                "ok": False,
                "slots": [],
                "message": f"Este profissional não atende nesse dia. Atende: {', '.join(days) or '—'}.",
                "days": days,
            }
        )

    slots = _get_slots(prof, date_obj, service=service)
    if not slots:
        return JsonResponse({"ok": False, "slots": [], "message": "Sem horários disponíveis."})
    return JsonResponse({"ok": True, "slots": slots, "message": "", "days": professional_weekdays_labels(prof)})


def healthcheck_view(request):
    from django.db import connection

    status = "ok"
    db_ok = True
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT 1")
            cursor.fetchone()
    except Exception:
        status = "error"
        db_ok = False

    return JsonResponse({"status": status, "db": db_ok})


def test_duralux_view(request):
    return render(request, "core/test_duralux.html")


def test_fullcalendar(request):
    return render(request, "test_fullcalendar.html")


def backoffice_cancel_appointment_view(request, appointment_id):
    if not can_view_all_calendar(request.user):
        return HttpResponseForbidden("Acesso apenas para receção/admin.")

    appt = get_object_or_404(Appointment, id=appointment_id)
    if appt.status in {Appointment.STATUS_COMPLETED, Appointment.STATUS_IN_DEBT, Appointment.STATUS_NO_SHOW}:
        return HttpResponseForbidden("Não podes cancelar uma marcação concluída, em dívida ou em falta.")

    status_field = Appointment._meta.get_field("status")
    choices = [c[0] for c in (status_field.choices or [])]
    old_status = appt.status
    if "cancelled" in choices:
        appt.status = "cancelled"
        appt.save(update_fields=["status"])
    else:
        appt.delete()

    if old_status != getattr(appt, "status", old_status):
        log_appt(
            AppointmentLog.ACTION_CANCELLED,
            appt,
            request.user,
            old_status=old_status,
            new_status=getattr(appt, "status", None),
            request=request,
        )
        settings_obj = clinic_settings()
        if settings_obj.notify_client_on_clinic_changes:
            client_email = (appt.client.email or "").strip()
            if client_email:
                send_templated_email(
                    client_email,
                    f"Marcação cancelada — {settings_obj.clinic_name}",
                    "emails/appointment_changed_by_clinic.html",
                    "emails/appointment_changed_by_clinic.txt",
                    {
                        "client_name": appt.client.get_full_name() or appt.client.username,
                        "change_type": "cancelled",
                        "old_date": appt.date,
                        "old_time": appt.time,
                        "new_date": "",
                        "new_time": "",
                        "service_name": appt.service.name if appt.service else "-",
                        "professional_name": appt.professional.user.get_full_name() or appt.professional.user.username,
                        "reason": "",
                        "manage_url": request.build_absolute_uri(reverse("my_appointments")),
                    },
                    event="cancel_clinic",
                )
            else:
                log_email_skip("cancel_clinic", "Marcação cancelada", "Cliente sem email", "")

    messages.success(request, "Marcação cancelada.")
    return_to = _safe_return_to(request, request.GET.get("return_to"))
    if return_to:
        return redirect(return_to)
    return redirect("backoffice_agenda")


def backoffice_complete_appointment_view(request, appointment_id):
    if not can_view_all_calendar(request.user):
        return HttpResponseForbidden("Acesso apenas para receção/admin.")

    appt = get_object_or_404(Appointment, id=appointment_id)
    return_to = _safe_return_to(request, request.GET.get("return_to"))
    params = {}
    if return_to:
        params["return_to"] = return_to
    target = reverse("professional_appointment_detail", args=[appt.id])
    if params:
        return redirect(f"{target}?{urlencode(params)}")
    return redirect(target)


@dataclass
class AgendaItem:
    kind: str
    appointment_id: int | None
    date: datetime.date
    time: datetime.time
    service_name: str
    professional_name: str
    client_label: str
    consumptions_label: str
    status_label: str
    status_raw: str
    price_label: str
    open_url: str
    cancel_url: str | None
    complete_url: str | None
    reschedule_url: str | None


def backoffice_professionals_list_view(request):
    q = (request.GET.get("q") or "").strip()
    employment = (request.GET.get("employment") or "").strip()
    per_page = request.GET.get("per_page") or "5"
    try:
        per_page = int(per_page)
    except (TypeError, ValueError):
        per_page = 5
    if per_page not in (5, 10, 15, 25, 50):
        per_page = 5

    qs = Professional.objects.select_related("user").prefetch_related("services").order_by(
        "user__username"
    )
    if q:
        qs = apply_terms_filter(
            qs,
            q,
            [
                "user__first_name__icontains",
                "user__last_name__icontains",
                "user__username__icontains",
                "speciality__icontains",
            ],
        )
    if employment == "independent":
        qs = qs.filter(is_independent=True)
    elif employment == "employee":
        qs = qs.filter(is_independent=False)

    paginator = Paginator(qs, per_page)
    page_obj = paginator.get_page(request.GET.get("page") or 1)

    query_params = {}
    if q:
        query_params["q"] = q
    if employment:
        query_params["employment"] = employment
    if per_page:
        query_params["per_page"] = per_page

    return render(
        request,
        "backoffice/professionals_list.html",
        {
            "professionals": page_obj.object_list,
            "page_obj": page_obj,
            "paginator": paginator,
            "q": q,
            "employment": employment,
            "per_page": per_page,
            "query_prefix": urlencode(query_params),
            "return_to": request.get_full_path(),
        },
    )


def _send_professional_welcome_email(request, professional):
    user = professional.user
    if not user or not user.email:
        log_email_skip("professional_welcome", "Acesso à plataforma", "Profissional sem email", "")
        return False, "Profissional sem email."

    login_url = request.build_absolute_uri(reverse("login"))
    context = {
        "professional": professional,
        "user": user,
        "login_url": login_url,
    }
    sent = send_templated_email(
        to_email=user.email,
        subject="Acesso à plataforma de marcações Fisio-UP",
        template_html="emails/professional_welcome.html",
        template_txt="emails/professional_welcome.txt",
        context=context,
        event="professional_welcome",
    )
    if sent:
        return True, "Email enviado."
    return False, "Falha ao enviar email."


def backoffice_professional_create_view(request):
    return_to = _safe_return_to(request, request.POST.get("return_to") or request.GET.get("return_to"))
    if request.method == "POST":
        form = BackofficeProfessionalForm(request.POST, request.FILES)
        form.fields.pop("user", None)
        if form.is_valid():
            professional = form.save()
            send_email = bool(request.POST.get("send_welcome_email"))
            log_audit_event(
                category="professionals",
                action="created",
                request=request,
                actor=request.user,
                instance=professional,
                source="backoffice_professional_create",
                message="Profissional criado no backoffice.",
                after=snapshot_instance(
                    professional,
                    fields=[
                        "user_id",
                        "speciality",
                        "gender",
                        "phone",
                        "is_independent",
                        "subcontract_percentage",
                        "hourly_rate",
                    ],
                ),
                metadata={
                    "services": list(professional.services.values_list("id", flat=True)),
                    "send_welcome_email": send_email,
                },
            )
            if send_email:
                ok, msg = _send_professional_welcome_email(request, professional)
                if ok:
                    messages.success(request, msg)
                else:
                    messages.warning(request, msg)
            messages.success(request, "Profissional criado com sucesso.")
            if return_to:
                return redirect(return_to)
            return redirect("backoffice_professionals")
    else:
        form = BackofficeProfessionalForm()
    return render(
        request,
        "backoffice/professional_form.html",
        {"form": form, "title": "Novo profissional", "return_to": return_to, "is_edit": False},
    )


def backoffice_professional_edit_view(request, professional_id):
    professional = get_object_or_404(Professional, id=professional_id)
    return_to = _safe_return_to(request, request.POST.get("return_to") or request.GET.get("return_to"))
    active_tab = "profile"
    password_form = SetPasswordForm(user=professional.user)
    for field in password_form.fields.values():
        field.widget.attrs.setdefault("class", "form-control")
    if request.method == "POST":
        if request.POST.get("action") == "set_password":
            active_tab = "password"
            password_form = SetPasswordForm(user=professional.user, data=request.POST)
            for field in password_form.fields.values():
                field.widget.attrs.setdefault("class", "form-control")
            if password_form.is_valid():
                password_form.save()
                log_audit_event(
                    category="professionals",
                    action="password_changed",
                    request=request,
                    actor=request.user,
                    instance=professional,
                    source="backoffice_professional_edit",
                    message="Password do profissional atualizada.",
                )
                messages.success(request, "Password atualizada com sucesso.")
                redirect_url = reverse("backoffice_professional_edit", args=[professional.id])
                if return_to:
                    redirect_url = f"{redirect_url}?{urlencode({'return_to': return_to})}"
                return redirect(redirect_url)
            form = BackofficeProfessionalForm(instance=professional)
        else:
            form = BackofficeProfessionalForm(request.POST, request.FILES, instance=professional)
            if form.is_valid():
                before = snapshot_instance(
                    professional,
                    fields=[
                        "user_id",
                        "speciality",
                        "gender",
                        "phone",
                        "is_independent",
                        "subcontract_percentage",
                        "hourly_rate",
                    ],
                )
                before_services = list(professional.services.values_list("id", flat=True))
                form.save()
                professional.refresh_from_db()
                log_audit_event(
                    category="professionals",
                    action="updated",
                    request=request,
                    actor=request.user,
                    instance=professional,
                    source="backoffice_professional_edit",
                    message="Profissional atualizado.",
                    before=before,
                    after=snapshot_instance(
                        professional,
                        fields=[
                            "user_id",
                            "speciality",
                            "gender",
                            "phone",
                            "is_independent",
                            "subcontract_percentage",
                            "hourly_rate",
                        ],
                    ),
                    metadata={
                        "services_before": before_services,
                        "services_after": list(professional.services.values_list("id", flat=True)),
                    },
                )
                messages.success(request, "Profissional atualizado.")
                if return_to:
                    return redirect(return_to)
                return redirect("backoffice_professionals")
    else:
        form = BackofficeProfessionalForm(instance=professional)
    return render(
        request,
        "backoffice/professional_form.html",
        {
            "form": form,
            "title": "Editar profissional",
            "return_to": return_to,
            "is_edit": True,
            "password_form": password_form,
            "active_tab": active_tab,
        },
    )


@backoffice_required
def backoffice_weekly_schedules_list_view(request):
    if not is_admin_role(request.user):
        return HttpResponseForbidden("Acesso reservado a administradores.")

    professionals = Professional.objects.select_related("user").order_by("user__username")
    schedules = WeeklySchedule.objects.select_related("professional").all()
    schedule_map = {s.professional_id: s for s in schedules}

    block_counts = {
        row["weekly_schedule__professional_id"]: row["total"]
        for row in WeeklyWorkingBlock.objects.values("weekly_schedule__professional_id").annotate(total=Count("id"))
    }

    rows = []
    for prof in professionals:
        schedule = schedule_map.get(prof.id)
        rows.append(
            {
                "professional": prof,
                "schedule": schedule,
                "has_schedule": bool(schedule),
                "is_active": bool(schedule and schedule.is_active),
                "block_count": block_counts.get(prof.id, 0),
            }
        )

    return render(
        request,
        "backoffice/weekly_schedules_list.html",
        {
            "rows": rows,
            "return_to": request.get_full_path(),
        },
    )


@backoffice_required
def backoffice_weekly_schedule_edit_view(request, professional_id):
    if not is_admin_role(request.user):
        return HttpResponseForbidden("Acesso reservado a administradores.")

    professional = get_object_or_404(Professional, id=professional_id)
    professionals = Professional.objects.select_related("user").order_by("user__username")
    schedule, _ = WeeklySchedule.objects.get_or_create(
        professional=professional,
        defaults={"timezone": "Europe/Lisbon", "is_active": True},
    )

    if request.method == "POST":
        before = _weekly_schedule_snapshot(schedule)
        schedule_form = WeeklyScheduleForm(request.POST, instance=schedule)
        work_formset = WeeklyWorkingBlockFormSet(request.POST, instance=schedule, prefix="work")
        break_formset = WeeklyBreakBlockFormSet(request.POST, instance=schedule, prefix="break")
        if schedule_form.is_valid() and work_formset.is_valid() and break_formset.is_valid():
            schedule_form.save()
            work_formset.save()
            break_formset.save()
            log_audit_event(
                category="weekly_schedule",
                action="update",
                request=request,
                instance=schedule,
                source="backoffice_weekly_schedule",
                message="Horário semanal atualizado.",
                before=before,
                after=_weekly_schedule_snapshot(schedule),
                metadata={
                    "professional_id": professional.id,
                    "working_blocks_count": schedule.blocks.count(),
                    "break_blocks_count": schedule.breaks.count(),
                },
            )
            messages.success(request, "Horário semanal atualizado.")
            return redirect("backoffice_weekly_schedules")
    else:
        schedule_form = WeeklyScheduleForm(instance=schedule)
        work_formset = WeeklyWorkingBlockFormSet(instance=schedule, prefix="work")
        break_formset = WeeklyBreakBlockFormSet(instance=schedule, prefix="break")

    weekdays = [
        (0, "Segunda-feira"),
        (1, "Terça-feira"),
        (2, "Quarta-feira"),
        (3, "Quinta-feira"),
        (4, "Sexta-feira"),
        (5, "Sábado"),
        (6, "Domingo"),
    ]

    def _weekday_from_form(form):
        if form.instance and getattr(form.instance, "weekday", None) is not None:
            return form.instance.weekday
        raw = form.data.get(f"{form.prefix}-weekday")
        if raw not in (None, ""):
            try:
                return int(raw)
            except ValueError:
                return None
        initial = form.initial.get("weekday")
        if initial is not None:
            return initial
        return None

    work_by_day = {day: [] for day, _ in weekdays}
    for form in work_formset.forms:
        day = _weekday_from_form(form)
        if day in work_by_day:
            work_by_day[day].append(form)

    break_by_day = {day: [] for day, _ in weekdays}
    for form in break_formset.forms:
        day = _weekday_from_form(form)
        if day in break_by_day:
            break_by_day[day].append(form)

    day_rows = [
        {
            "day": day,
            "label": label,
            "work_forms": work_by_day.get(day, []),
            "break_forms": break_by_day.get(day, []),
        }
        for day, label in weekdays
    ]

    return render(
        request,
        "backoffice/weekly_schedule_edit.html",
        {
            "professional": professional,
            "professionals": professionals,
            "schedule_form": schedule_form,
            "work_formset": work_formset,
            "break_formset": break_formset,
            "day_rows": day_rows,
            "return_to": request.get_full_path(),
        },
    )


def backoffice_partners_list_view(request):
    q = (request.GET.get("q") or "").strip()
    per_page = request.GET.get("per_page") or "5"
    try:
        per_page = int(per_page)
    except (TypeError, ValueError):
        per_page = 5
    if per_page not in (5, 10, 15, 25, 50):
        per_page = 5

    qs = Partner.objects.all().order_by("name")
    if q:
        qs = apply_terms_filter(qs, q, ["name__icontains"])

    paginator = Paginator(qs, per_page)
    page_obj = paginator.get_page(request.GET.get("page") or 1)

    return render(
        request,
        "backoffice/partners_list.html",
        {
            "partners": page_obj.object_list,
            "page_obj": page_obj,
            "paginator": paginator,
            "q": q,
            "per_page": per_page,
            "return_to": request.get_full_path(),
        },
    )


def backoffice_partner_create_view(request):
    return_to = _safe_return_to(request, request.POST.get("return_to") or request.GET.get("return_to"))
    if request.method == "POST":
        form = BackofficePartnerForm(request.POST, request.FILES)
        if form.is_valid():
            partner_obj = form.save(commit=False)
            partner_obj.discount_type = "none"
            partner_obj.discount_percent = None
            partner_obj.discount_amount = None
            partner_obj.discount_label = ""
            partner_obj.save()
            log_audit_event(
                category="partner",
                action="create",
                request=request,
                instance=partner_obj,
                source="backoffice_partners",
                message="Parceria criada.",
                after=_partner_snapshot(partner_obj),
            )
            messages.success(request, "Parceria criada com sucesso.")
            if return_to:
                return redirect(return_to)
            return redirect("backoffice_partners")
    else:
        form = BackofficePartnerForm()
    return render(
        request,
        "backoffice/partner_form.html",
        {"form": form, "title": "Nova parceria", "return_to": return_to},
    )


def backoffice_partner_edit_view(request, partner_id):
    partner = get_object_or_404(Partner, id=partner_id)
    return_to = _safe_return_to(request, request.POST.get("return_to") or request.GET.get("return_to"))
    if request.method == "POST":
        before = _partner_snapshot(partner)
        form = BackofficePartnerForm(request.POST, request.FILES, instance=partner)
        if form.is_valid():
            partner_obj = form.save(commit=False)
            partner_obj.discount_type = "none"
            partner_obj.discount_percent = None
            partner_obj.discount_amount = None
            partner_obj.discount_label = ""
            partner_obj.save()
            log_audit_event(
                category="partner",
                action="update",
                request=request,
                instance=partner_obj,
                source="backoffice_partners",
                message="Parceria atualizada.",
                before=before,
                after=_partner_snapshot(partner_obj),
            )
            messages.success(request, "Parceria atualizada.")
            if return_to:
                return redirect(return_to)
            return redirect("backoffice_partners")
    else:
        form = BackofficePartnerForm(instance=partner)
    return render(
        request,
        "backoffice/partner_form.html",
        {"form": form, "title": "Editar parceria", "return_to": return_to},
    )


def backoffice_partner_prices_view(request):
    partners = Partner.objects.order_by("name")
    partner_id = request.GET.get("partner_id") or request.POST.get("partner_id") or ""
    selected_partner = Partner.objects.filter(id=partner_id).first() if partner_id else None
    services = Service.objects.order_by("name")

    if request.method == "POST" and selected_partner:
        before_prices = _partner_prices_snapshot(selected_partner)
        existing_map = {
            psp.service_id: psp
            for psp in PartnerServicePrice.objects.filter(partner=selected_partner)
        }

        def _to_decimal(val):
            try:
                return Decimal(str(val).replace(",", "."))
            except Exception:
                return None

        for service in services:
            prefix = f"service_{service.id}"
            effect_mode = (request.POST.get(f"{prefix}_effect_mode") or "none").strip().lower()
            if effect_mode not in {"none", "discount", "final_price"}:
                effect_mode = "none"
            pricing_mode = request.POST.get(f"{prefix}_pricing_mode") or "single"
            if pricing_mode not in {"single", "first_followup"}:
                pricing_mode = "single"
            discount_type = (request.POST.get(f"{prefix}_discount_type") or "percent").strip().lower()
            if discount_type not in {"percent", "fixed"}:
                discount_type = "percent"
            price = (request.POST.get(f"{prefix}_price") or "").strip()
            price_first = (request.POST.get(f"{prefix}_price_first") or "").strip()
            price_followup = (request.POST.get(f"{prefix}_price_followup") or "").strip()
            discount_value_raw = (request.POST.get(f"{prefix}_discount") or "").strip()

            price_val = _to_decimal(price)
            price_first_val = _to_decimal(price_first)
            price_followup_val = _to_decimal(price_followup)
            discount_val = _to_decimal(discount_value_raw) if discount_value_raw else None
            obj = existing_map.get(service.id)
            has_row_input = bool(price or price_first or price_followup or discount_value_raw)

            if effect_mode == "none":
                if obj and obj.is_enabled:
                    obj.is_enabled = False
                    obj.full_clean()
                    obj.save(update_fields=["is_enabled", "updated_at"])
                continue

            if not obj and not has_row_input:
                continue

            obj, _ = PartnerServicePrice.objects.get_or_create(
                partner=selected_partner,
                service=service,
                defaults={
                    "price": Decimal(service.price or 0),
                    "pricing_mode": service.pricing_mode or "single",
                    "price_first": service.price_first,
                    "price_followup": service.price_followup,
                    "discount_type": "none",
                    "discount_percent": None,
                    "discount_amount": None,
                    "is_enabled": True,
                },
            )
            existing_map[service.id] = obj
            obj.is_enabled = True
            if effect_mode == "discount":
                obj.pricing_mode = service.pricing_mode or "single"
                obj.price = Decimal(service.price or 0)
                obj.price_first = service.price_first
                obj.price_followup = service.price_followup
                obj.discount_type = discount_type
                if discount_type == "percent":
                    obj.discount_percent = discount_val if discount_val is not None else Decimal("0.00")
                    obj.discount_amount = None
                else:
                    obj.discount_amount = discount_val if discount_val is not None else Decimal("0.00")
                    obj.discount_percent = None
            else:
                obj.pricing_mode = pricing_mode
                if pricing_mode == "single":
                    if price_val is not None:
                        obj.price = price_val
                    elif obj.price is None:
                        obj.price = Decimal(service.price or 0)
                    obj.price_first = None
                    obj.price_followup = None
                else:
                    if price_first_val is not None:
                        obj.price_first = price_first_val
                    elif obj.price_first is None:
                        obj.price_first = Decimal(service.price_first or service.price or 0)
                    if price_followup_val is not None:
                        obj.price_followup = price_followup_val
                    elif obj.price_followup is None:
                        obj.price_followup = Decimal(service.price_followup or service.price or 0)
                    if obj.price is None:
                        obj.price = Decimal(service.price or 0)
                obj.discount_type = "none"
                obj.discount_percent = None
                obj.discount_amount = None

            obj.full_clean()
            obj.save()

        # Recalcula sempre todas as marcações futuras desta parceria.
        # Evita deixar marcações pendentes com preço antigo quando há
        # alterações de modo/valor em múltiplos serviços.
        updated_upcoming = recalculate_partner_upcoming_appointments(selected_partner)

        if updated_upcoming:
            messages.success(
                request,
                f"Preços atualizados. {updated_upcoming} marcações futuras foram recalculadas.",
            )
        else:
            messages.success(request, "Preços atualizados.")
        log_audit_event(
            category="partner_service_price",
            action="bulk_update",
            request=request,
            instance=selected_partner,
            source="backoffice_partner_prices",
            message="Preços da parceria atualizados.",
            before={"prices": before_prices},
            after={"prices": _partner_prices_snapshot(selected_partner)},
            metadata={
                "partner_id": selected_partner.id,
                "updated_upcoming_appointments": updated_upcoming,
            },
        )
        return redirect(f"{reverse('backoffice_partner_prices')}?partner_id={selected_partner.id}")

    service_rows = []
    if selected_partner:
        price_map = {
            psp.service_id: psp
            for psp in PartnerServicePrice.objects.filter(partner=selected_partner)
        }
        for service in services:
            psp = price_map.get(service.id)
            if not psp:
                effect_mode = "none"
            elif not psp.is_enabled:
                effect_mode = "none"
            elif psp.discount_type != "none":
                effect_mode = "discount"
            else:
                effect_mode = "final_price"
            service_rows.append(
                {
                    "service": service,
                    "psp": psp,
                    "effect_mode": effect_mode,
                    "discount_type": psp.discount_type if psp and psp.discount_type in {"percent", "fixed"} else "percent",
                }
            )

    return render(
        request,
        "backoffice/partner_prices.html",
        {
            "partners": partners,
            "selected_partner": selected_partner,
            "services": services,
            "service_rows": service_rows,
        },
    )


def _moloni_redirect_uri(request):
    configured = (getattr(settings, "MOLONI_REDIRECT_URI", "") or "").strip()
    if configured:
        return configured
    return request.build_absolute_uri(reverse("backoffice_moloni_callback"))


@backoffice_required
def backoffice_settings_moloni_view(request):
    if not is_admin_role(request.user):
        return HttpResponseForbidden("Acesso reservado a administradores.")

    integ = MoloniIntegration.get_solo()
    defaults_initial = {
        "payment_method_id": integ.customer_payment_method_id,
        "document_type_id": integ.customer_document_type_id,
        "language_id": integ.customer_language_id,
        "maturity_date_id": integ.customer_maturity_date_id,
        "country_id": integ.customer_country_id,
        "delivery_method_id": integ.customer_delivery_method_id,
    }
    defaults_form = MoloniCustomerDefaultsForm(initial=defaults_initial)
    if request.method == "POST" and (request.POST.get("action") or "").strip() == "save_customer_defaults":
        defaults_form = MoloniCustomerDefaultsForm(request.POST)
        if defaults_form.is_valid():
            before = snapshot_instance(
                integ,
                fields=[
                    "customer_payment_method_id",
                    "customer_document_type_id",
                    "customer_language_id",
                    "customer_maturity_date_id",
                    "customer_country_id",
                    "customer_delivery_method_id",
                ],
            )
            integ.customer_payment_method_id = defaults_form.cleaned_data["payment_method_id"]
            integ.customer_document_type_id = defaults_form.cleaned_data["document_type_id"]
            integ.customer_language_id = defaults_form.cleaned_data["language_id"]
            integ.customer_maturity_date_id = defaults_form.cleaned_data["maturity_date_id"]
            integ.customer_country_id = defaults_form.cleaned_data["country_id"]
            integ.customer_delivery_method_id = defaults_form.cleaned_data.get("delivery_method_id")
            integ.save(
                update_fields=[
                    "customer_payment_method_id",
                    "customer_document_type_id",
                    "customer_language_id",
                    "customer_maturity_date_id",
                    "customer_country_id",
                    "customer_delivery_method_id",
                    "updated_at",
                ]
            )
            log_audit_event(
                category="integrations",
                action="moloni_customer_defaults_saved",
                request=request,
                actor=request.user,
                instance=integ,
                source="backoffice_settings_moloni",
                message="Defaults de clientes Moloni guardados.",
                before=before,
                after=snapshot_instance(
                    integ,
                    fields=[
                        "customer_payment_method_id",
                        "customer_document_type_id",
                        "customer_language_id",
                        "customer_maturity_date_id",
                        "customer_country_id",
                        "customer_delivery_method_id",
                    ],
                ),
            )
            messages.success(request, "Defaults de clientes Moloni guardados com sucesso.")
            return redirect("backoffice_settings_moloni")

    configured = moloni_service.is_configured()
    connected = bool(integ.refresh_token)
    callback_url = _moloni_redirect_uri(request)
    companies = []
    companies_error = ""
    defaults_suggestions = None
    defaults_suggestions_error = ""
    if connected and not moloni_service.get_company_id():
        try:
            companies = moloni_service.list_companies()
        except moloni_service.MoloniError as exc:
            companies_error = str(exc)
    elif connected and moloni_service.get_company_id():
        try:
            defaults_suggestions = moloni_service.get_customer_defaults_suggestions()
        except moloni_service.MoloniError as exc:
            defaults_suggestions_error = str(exc)
    defaults_status = moloni_service.get_customer_defaults_status()
    context = {
        "moloni_integration": integ,
        "moloni_configured": configured,
        "moloni_connected": connected,
        "moloni_ready": bool(connected and moloni_service.get_company_id()),
        "moloni_callback_url": callback_url,
        "moloni_client_id": getattr(settings, "MOLONI_CLIENT_ID", ""),
        "moloni_company_id": moloni_service.get_company_id(),
        "moloni_company_name": moloni_service.get_company_name(),
        "moloni_company_id_from_env": getattr(settings, "MOLONI_COMPANY_ID", ""),
        "moloni_base_url": getattr(settings, "MOLONI_BASE_URL", ""),
        "moloni_companies": companies,
        "moloni_companies_error": companies_error,
        "moloni_defaults_form": defaults_form,
        "moloni_defaults_ready": defaults_status["ready"],
        "moloni_defaults_missing": defaults_status["missing"],
        "moloni_defaults_suggestions": defaults_suggestions,
        "moloni_defaults_suggestions_error": defaults_suggestions_error,
    }
    return render(request, "backoffice/settings_moloni.html", context)


@backoffice_required
def backoffice_moloni_connect_view(request):
    if not is_admin_role(request.user):
        return HttpResponseForbidden("Acesso reservado a administradores.")
    if not moloni_service.is_configured():
        messages.error(request, "Configuração Moloni incompleta. Define as variáveis da API primeiro.")
        return redirect("backoffice_settings_moloni")

    authorize_url = moloni_service.build_authorize_url(
        redirect_uri=_moloni_redirect_uri(request),
    )
    return redirect(authorize_url)


@backoffice_required
def backoffice_moloni_callback_view(request):
    if not is_admin_role(request.user):
        return HttpResponseForbidden("Acesso reservado a administradores.")

    error = (request.GET.get("error") or "").strip()
    code = (request.GET.get("code") or "").strip()

    if error:
        messages.error(request, f"Moloni devolveu erro na autorização: {error}")
        return redirect("backoffice_settings_moloni")
    if not code:
        messages.error(request, "Moloni não devolveu o código de autorização.")
        return redirect("backoffice_settings_moloni")

    try:
        data = moloni_service.exchange_authorization_code(
            code=code,
            redirect_uri=_moloni_redirect_uri(request),
        )
        integ = moloni_service.store_tokens(data)
        company_data = {}
        company_message = "Integração Moloni ligada com sucesso."
        try:
            company_data = moloni_service.discover_company()
            company_message = f"Integração Moloni ligada com sucesso à empresa {company_data['company_name'] or company_data['company_id']}."
        except moloni_service.MoloniError:
            company_message = "Integração Moloni ligada. Escolhe agora a empresa manualmente no painel."
    except moloni_service.MoloniError as exc:
        messages.error(request, str(exc))
        return redirect("backoffice_settings_moloni")

    log_audit_event(
        category="integrations",
        action="moloni_connected",
        request=request,
        actor=request.user,
        instance=integ,
        source="backoffice_settings_moloni",
        message="Integração Moloni ligada com sucesso.",
        after={
            **snapshot_instance(integ, fields=["expires_at", "last_sync_at"]),
            **company_data,
        },
    )
    if company_data:
        messages.success(request, company_message)
    else:
        messages.warning(request, company_message)
    return redirect("backoffice_settings_moloni")


@require_POST
@backoffice_required
def backoffice_moloni_select_company_view(request):
    if not is_admin_role(request.user):
        return HttpResponseForbidden("Acesso reservado a administradores.")

    company_id = (request.POST.get("company_id") or "").strip()
    if not company_id:
        messages.error(request, "Seleciona uma empresa Moloni.")
        return redirect("backoffice_settings_moloni")

    try:
        companies = moloni_service.list_companies()
    except moloni_service.MoloniError as exc:
        messages.error(request, f"Não foi possível listar as empresas Moloni: {exc}")
        return redirect("backoffice_settings_moloni")

    selected = next((company for company in companies if company["company_id"] == company_id), None)
    if not selected:
        messages.error(request, "A empresa selecionada não está disponível para esta conta Moloni.")
        return redirect("backoffice_settings_moloni")

    integ = moloni_service.store_company(selected["company_id"], selected["company_name"])
    log_audit_event(
        category="integrations",
        action="moloni_company_selected",
        request=request,
        actor=request.user,
        instance=integ,
        source="backoffice_settings_moloni",
        message="Empresa Moloni selecionada manualmente.",
        after=selected,
    )
    messages.success(request, f"Empresa Moloni definida: {selected['company_name']}.")
    return redirect("backoffice_settings_moloni")


@require_POST
@backoffice_required
def backoffice_moloni_test_view(request):
    if not is_admin_role(request.user):
        return HttpResponseForbidden("Acesso reservado a administradores.")

    try:
        result = moloni_service.test_connection()
    except moloni_service.MoloniError as exc:
        messages.error(request, f"Teste Moloni falhou: {exc}")
        return redirect("backoffice_settings_moloni")

    log_audit_event(
        category="integrations",
        action="moloni_tested",
        request=request,
        actor=request.user,
        instance=MoloniIntegration.get_solo(),
        source="backoffice_settings_moloni",
        message="Ligação Moloni testada com sucesso.",
        after=result,
    )
    messages.success(request, "Ligação Moloni testada com sucesso.")
    return redirect("backoffice_settings_moloni")


@require_POST
@backoffice_required
def backoffice_moloni_sync_customers_view(request):
    if not is_admin_role(request.user):
        return HttpResponseForbidden("Acesso reservado a administradores.")

    sync_mode = (request.POST.get("sync_mode") or "incremental").strip().lower()
    full = sync_mode == "full"
    try:
        result = moloni_sync_customers(full=full)
    except moloni_service.MoloniError as exc:
        messages.error(request, f"Sincronização Moloni falhou: {exc}")
        return redirect("backoffice_settings_moloni")

    log_audit_event(
        category="integrations",
        action="moloni_customers_synced",
        request=request,
        actor=request.user,
        instance=MoloniIntegration.get_solo(),
        source="backoffice_settings_moloni",
        message="Clientes Moloni sincronizados.",
        after=result,
    )
    messages.success(
        request,
        f"Sincronização {result['mode']} concluída. Criados: {result['created']}, atualizados: {result['updated']}, ignorados: {result['skipped']}, erros: {result['errors']}.",
    )
    return redirect("backoffice_settings_moloni")


@require_POST
@backoffice_required
def backoffice_moloni_run_reconciliation_view(request):
    if not is_admin_role(request.user):
        return HttpResponseForbidden("Acesso reservado a administradores.")

    sync_mode = (request.POST.get("sync_mode") or "incremental").strip().lower()
    full = sync_mode == "full"
    try:
        result = moloni_run_bidirectional_reconciliation(full=full)
    except moloni_service.MoloniError as exc:
        messages.error(request, f"Reconciliação Moloni falhou: {exc}")
        return redirect("backoffice_moloni_reconciliation")

    log_audit_event(
        category="integrations",
        action="moloni_reconciliation_run",
        request=request,
        actor=request.user,
        instance=MoloniIntegration.get_solo(),
        source="backoffice_moloni_reconciliation",
        message="Reconciliação Moloni executada.",
        after=result,
    )
    messages.success(request, "Reconciliação Moloni executada com sucesso.")
    return redirect("backoffice_moloni_reconciliation")


@backoffice_required
def backoffice_moloni_reconciliation_view(request):
    if not is_admin_role(request.user):
        return HttpResponseForbidden("Acesso reservado a administradores.")

    integ = MoloniIntegration.get_solo()
    report = None
    report_error = ""
    if moloni_service.is_configured() and integ.refresh_token and moloni_service.get_company_id():
        try:
            report = moloni_build_reconciliation_report(limit=50)
        except moloni_service.MoloniError as exc:
            report_error = str(exc)

    return render(
        request,
        "backoffice/settings_moloni_reconciliation.html",
        {
            "moloni_integration": integ,
            "moloni_ready": bool(integ.refresh_token and moloni_service.get_company_id()),
            "moloni_company_id": moloni_service.get_company_id(),
            "moloni_company_name": moloni_service.get_company_name(),
            "report": report,
            "report_error": report_error,
        },
    )


@require_POST
@backoffice_required
def backoffice_moloni_apply_remote_view(request, client_id):
    if not is_admin_role(request.user):
        return HttpResponseForbidden("Acesso reservado a administradores.")

    profile = get_object_or_404(ClientProfile.objects.select_related("user"), id=client_id)
    try:
        result = moloni_apply_remote_customer_to_profile(profile)
    except moloni_service.MoloniError as exc:
        messages.error(request, f"Não foi possível atualizar a app com dados da Moloni: {exc}")
        return redirect("backoffice_moloni_reconciliation")

    log_audit_event(
        category="integrations",
        action="moloni_apply_remote_to_app",
        request=request,
        actor=request.user,
        instance=profile,
        source="backoffice_moloni_reconciliation",
        message="Dados do cliente atualizados na app a partir da Moloni.",
        after=result,
    )
    messages.success(request, "Dados do cliente atualizados na app a partir da Moloni.")
    return redirect("backoffice_moloni_reconciliation")


@require_POST
@backoffice_required
def backoffice_moloni_disconnect_view(request):
    if not is_admin_role(request.user):
        return HttpResponseForbidden("Acesso reservado a administradores.")

    integ = moloni_service.disconnect()
    log_audit_event(
        category="integrations",
        action="moloni_disconnected",
        request=request,
        actor=request.user,
        instance=integ,
        source="backoffice_settings_moloni",
        message="Integração Moloni desligada.",
        before={"connected": True},
        after={"connected": False},
    )
    messages.success(request, "Integração Moloni desligada.")
    return redirect("backoffice_settings_moloni")


@backoffice_required
def backoffice_settings_email_view(request):
    if not is_admin_role(request.user):
        return HttpResponseForbidden("Acesso reservado a administradores.")

    settings_obj = ClinicSettings.get_solo()
    form = ClinicEmailSettingsForm(request.POST or None, instance=settings_obj)
    if request.method == "POST" and form.is_valid():
        before = snapshot_instance(settings_obj, fields=form.fields.keys())
        form.save()
        settings_obj.refresh_from_db()
        log_audit_event(
            category="settings",
            action="email_settings_updated",
            request=request,
            actor=request.user,
            instance=settings_obj,
            source="backoffice_settings_email",
            message="Definições de email atualizadas.",
            before=before,
            after=snapshot_instance(settings_obj, fields=form.fields.keys()),
        )
        messages.success(request, "Definições de email atualizadas.")
        return redirect("backoffice_settings_email")

    return render(
        request,
        "backoffice/settings_email.html",
        {"form": form},
    )


@backoffice_required
def backoffice_audit_logs_view(request):
    if not is_admin_role(request.user):
        return HttpResponseForbidden("Acesso reservado a administradores.")

    cleanup_old_audit_logs_if_needed()
    dashboard_data = _get_cached_audit_dashboard_data()
    qs = AuditLog.objects.select_related("actor", "content_type").order_by("-created_at", "-id")
    q = (request.GET.get("q") or "").strip()
    actor_email = (request.GET.get("actor_email") or "").strip()
    object_type = (request.GET.get("object_type") or "all").strip()
    category = (request.GET.get("category") or "all").strip()
    action = (request.GET.get("action") or "all").strip()
    actor_role = (request.GET.get("actor_role") or "all").strip()
    source = (request.GET.get("source") or "all").strip()
    date_from = (request.GET.get("date_from") or "").strip()
    date_to = (request.GET.get("date_to") or "").strip()
    try:
        per_page = int(request.GET.get("per_page") or 10)
    except (TypeError, ValueError):
        per_page = 10
    if per_page not in (10, 25, 50, 100):
        per_page = 10

    if q:
        qs = apply_terms_filter(
            qs,
            q,
            [
                "object_repr__icontains",
                "message__icontains",
                "actor_display__icontains",
                "actor_email__icontains",
                "request_path__icontains",
            ],
        )
    if actor_email:
        qs = qs.filter(actor_email__icontains=actor_email)
    if object_type != "all":
        if object_type == "none":
            qs = qs.filter(content_type__isnull=True)
        elif "." in object_type:
            app_label, model = object_type.split(".", 1)
            qs = qs.filter(content_type__app_label=app_label, content_type__model=model)
    if category != "all":
        qs = qs.filter(category=category)
    if action != "all":
        qs = qs.filter(action=action)
    if actor_role != "all":
        qs = qs.filter(actor_role=actor_role)
    if source != "all":
        qs = qs.filter(source=source)
    if date_from:
        qs = qs.filter(created_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(created_at__date__lte=date_to)

    if (request.GET.get("export") or "").strip().lower() == "csv":
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        timestamp = timezone.localtime().strftime("%Y%m%d-%H%M%S")
        response["Content-Disposition"] = f'attachment; filename="audit-log-{timestamp}.csv"'
        writer = csv.writer(response)
        writer.writerow([
            "created_at",
            "category",
            "action",
            "source",
            "actor_display",
            "actor_email",
            "actor_role",
            "object_type",
            "object_id",
            "object_repr",
            "message",
            "request_method",
            "request_path",
            "ip_address",
            "before",
            "after",
            "metadata",
        ])
        for log in qs.iterator():
            object_type_value = "none"
            if log.content_type_id:
                object_type_value = f"{log.content_type.app_label}.{log.content_type.model}"
            writer.writerow([
                timezone.localtime(log.created_at).strftime("%Y-%m-%d %H:%M:%S"),
                log.category,
                log.action,
                log.source,
                log.actor_display,
                log.actor_email,
                log.actor_role,
                object_type_value,
                log.object_id or "",
                log.object_repr,
                log.message,
                log.request_method,
                log.request_path,
                log.ip_address or "",
                json.dumps(log.before or {}, ensure_ascii=False, sort_keys=True),
                json.dumps(log.after or {}, ensure_ascii=False, sort_keys=True),
                json.dumps(log.metadata or {}, ensure_ascii=False, sort_keys=True),
            ])
        return response

    filter_params = request.GET.copy()
    filter_params.pop("page", None)
    filter_params.pop("export", None)
    filter_querystring = filter_params.urlencode()

    paginator = Paginator(qs, per_page)
    page_obj = paginator.get_page(request.GET.get("page") or 1)
    logs = list(page_obj.object_list)
    for log in logs:
        log.object_url = _audit_log_object_url(log)

    return render(
        request,
        "backoffice/audit_logs.html",
        {
            "logs": logs,
            "page_obj": page_obj,
            "q": q,
            "actor_email": actor_email,
            "object_type": object_type,
            "category": category,
            "action": action,
            "actor_role": actor_role,
            "source": source,
            "date_from": date_from,
            "date_to": date_to,
            "per_page": per_page,
            "filter_querystring": filter_querystring,
            "retention_days": getattr(settings, "AUDIT_LOG_RETENTION_DAYS", 365),
            **dashboard_data,
        },
    )


@backoffice_required
@require_GET
def backoffice_audit_log_detail_view(request, log_id):
    if not is_admin_role(request.user):
        return HttpResponseForbidden("Acesso reservado a administradores.")

    log = get_object_or_404(
        AuditLog.objects.select_related("actor", "content_type"),
        pk=log_id,
    )
    return JsonResponse(
        {
            "ok": True,
            "created_at": timezone.localtime(log.created_at).strftime("%d/%m/%Y %H:%M:%S"),
            "request_line": f"{log.request_method or '-'} {log.request_path or '-'}",
            "before": _audit_log_pretty_payload(log.before),
            "after": _audit_log_pretty_payload(log.after),
            "metadata": _audit_log_pretty_payload(log.metadata),
        }
    )


@backoffice_required
def backoffice_highlights_list_view(request):
    if not is_admin_role(request.user):
        return HttpResponseForbidden("Acesso reservado a administradores.")

    q = (request.GET.get("q") or "").strip()
    status = (request.GET.get("status") or "all").strip().lower()
    per_page = request.GET.get("per_page") or "6"
    try:
        per_page = int(per_page)
    except (TypeError, ValueError):
        per_page = 6
    if per_page not in (6, 12, 24):
        per_page = 6

    qs = ContentPost.objects.select_related("author").all().order_by("-is_featured", "-updated_at")
    if status in {"draft", "published"}:
        qs = qs.filter(status=status)
    else:
        status = "all"

    if q:
        qs = apply_terms_filter(
            qs,
            q,
            [
                "title__icontains",
                "excerpt__icontains",
                "body__icontains",
                "slug__icontains",
            ],
        )

    paginator = Paginator(qs, per_page)
    page_obj = paginator.get_page(request.GET.get("page") or 1)

    return render(
        request,
        "backoffice/highlights_list.html",
        {
            "posts": page_obj.object_list,
            "page_obj": page_obj,
            "q": q,
            "status": status,
            "per_page": per_page,
            "return_to": request.get_full_path(),
        },
    )


@backoffice_required
def backoffice_highlight_create_view(request):
    if not is_admin_role(request.user):
        return HttpResponseForbidden("Acesso reservado a administradores.")

    return_to = _safe_return_to(request, request.POST.get("return_to") or request.GET.get("return_to"))
    if request.method == "POST":
        form = BackofficeHighlightForm(request.POST, request.FILES)
        if form.is_valid():
            post = form.save(commit=False)
            post.author = request.user
            post.save()
            log_audit_event(
                category="content_post",
                action="create",
                request=request,
                instance=post,
                source="backoffice_highlights",
                message="Destaque criado no backoffice.",
                after=snapshot_instance(post),
            )
            messages.success(request, "Destaque criado com sucesso.")
            if return_to:
                return redirect(return_to)
            return redirect("backoffice_highlights")
    else:
        form = BackofficeHighlightForm()

    back_url = return_to or reverse("backoffice_highlights")
    return render(
        request,
        "backoffice/highlight_form.html",
        {
            "form": form,
            "title": "Novo destaque",
            "return_to": return_to,
            "back_url": back_url,
            "is_edit": False,
        },
    )


@backoffice_required
def backoffice_highlight_edit_view(request, post_id):
    if not is_admin_role(request.user):
        return HttpResponseForbidden("Acesso reservado a administradores.")

    post = get_object_or_404(ContentPost, id=post_id)
    return_to = _safe_return_to(request, request.POST.get("return_to") or request.GET.get("return_to"))
    if request.method == "POST":
        before = snapshot_instance(post)
        form = BackofficeHighlightForm(request.POST, request.FILES, instance=post)
        if form.is_valid():
            updated_post = form.save(commit=False)
            if not updated_post.author_id:
                updated_post.author = request.user
            updated_post.save()
            log_audit_event(
                category="content_post",
                action="update",
                request=request,
                instance=updated_post,
                source="backoffice_highlights",
                message="Destaque atualizado no backoffice.",
                before=before,
                after=snapshot_instance(updated_post),
            )
            messages.success(request, "Destaque atualizado.")
            if return_to:
                return redirect(return_to)
            return redirect("backoffice_highlights")
    else:
        form = BackofficeHighlightForm(instance=post)

    back_url = return_to or reverse("backoffice_highlights")
    return render(
        request,
        "backoffice/highlight_form.html",
        {
            "form": form,
            "title": "Editar destaque",
            "return_to": return_to,
            "back_url": back_url,
            "is_edit": True,
            "post": post,
        },
    )


def backoffice_clients_list_view(request):
    target_url = reverse("professional_clients")
    query_string = request.META.get("QUERY_STRING", "")
    if query_string:
        target_url = f"{target_url}?{query_string}"
    return redirect(target_url)


def backoffice_client_edit_view(request, client_id):
    target_url = reverse("prof_customer_edit", args=[client_id])
    query_string = request.META.get("QUERY_STRING", "")
    if query_string:
        target_url = f"{target_url}?{query_string}"
    return redirect(target_url)


def client_import_view(request):
    if not can_access_backoffice(request.user):
        return HttpResponseForbidden("Acesso apenas para backoffice.")

    def nif_is_valid(value: str) -> bool:
        value = "".join(ch for ch in (value or "") if ch.isdigit())
        if len(value) != 9:
            return False
        digits = [int(d) for d in value]
        total = sum(d * (9 - i) for i, d in enumerate(digits[:8]))
        check = 11 - (total % 11)
        check = 0 if check >= 10 else check
        return check == digits[8]

    selected_ids = set(request.session.get("import_selected_ids", []))
    batch_id = request.session.get("client_import_batch_id")

    def _norm(value: str) -> str:
        value = (value or "").strip().lower()
        value = "".join(
            ch for ch in unicodedata.normalize("NFKD", value) if not unicodedata.combining(ch)
        )
        value = value.replace(" ", "").replace("-", "").replace("_", "")
        return value

    def get_field(row, header_map, *names):
        for name in names:
            key = header_map.get(_norm(name))
            if key and key in row:
                return (row.get(key) or "").strip()
        return ""

    def parse_csv_to_batch(upload, validate_nif):
        raw = upload.read()
        content = ""
        for enc in ("utf-8-sig", "utf-8", "cp1252", "iso-8859-1"):
            try:
                content = raw.decode(enc)
                break
            except Exception:
                continue
        if not content:
            raise ValueError("Erro a ler o ficheiro. Usa UTF-8 ou ANSI (Windows-1252).")

        sample = content[:2048]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;")
        except Exception:
            dialect = csv.get_dialect("excel")

        lines = [line for line in content.splitlines() if line.strip()]
        if not lines:
            raise ValueError("CSV vazio.")

        start_idx = 0
        for i, line in enumerate(lines[:10]):
            if "nome" in line.lower() and ("contribuinte" in line.lower() or "nif" in line.lower()):
                start_idx = i
                break
        cleaned = "\n".join(lines[start_idx:])
        reader = csv.DictReader(io.StringIO(cleaned), dialect=dialect)

        header_map = {}
        if reader.fieldnames:
            for h in reader.fieldnames:
                header_map[_norm(h)] = h

        rows = [r for r in reader if any((v or "").strip() for v in r.values())]
        nif_counts = {}
        name_phone_counts = {}
        parsed = []
        for row in rows:
            full_name = get_field(row, header_map, "nome", "name", "full_name")
            phone = get_field(row, header_map, "telefone", "phone", "mobile", "telemovel")
            pair_key = build_client_name_phone_key(full_name, phone)
            nif_raw = get_field(row, header_map, "nif", "vat", "nif_cliente", "tax_id", "contribuinte")
            nif = "".join(ch for ch in nif_raw if ch.isdigit())
            if nif:
                nif_counts[nif] = nif_counts.get(nif, 0) + 1
            if pair_key:
                name_phone_counts[pair_key] = name_phone_counts.get(pair_key, 0) + 1
            parsed.append((row, full_name, phone, pair_key, nif_raw, nif))

        batch = ClientImportBatch.objects.create(
            uploaded_by=request.user,
            original_filename=upload.name,
            validate_nif=validate_nif,
        )
        bulk = []
        for idx, (row, full_name, phone, pair_key, nif_raw, nif) in enumerate(parsed, start=1):
            is_valid_nif = bool(nif) and nif_is_valid(nif)
            email = get_field(row, header_map, "email", "e-mail")
            address = get_field(row, header_map, "morada", "address", "address_line1")
            postal_code = get_field(row, header_map, "codigo postal", "codigo_postal", "postal_code", "cp", "zip")
            city = get_field(row, header_map, "localidade", "city")
            county = get_field(row, header_map, "concelho", "county")
            district = get_field(row, header_map, "distrito", "district")
            duplicate_in_file = (bool(nif) and nif_counts.get(nif, 0) > 1) or (
                bool(pair_key) and name_phone_counts.get(pair_key, 0) > 1
            )
            exists_in_db = (bool(nif) and ClientProfile.objects.filter(nif=nif).exists()) or (
                bool(pair_key) and bool(find_existing_client_by_name_phone(full_name, phone))
            )
            missing_email = not bool(email)
            bulk.append(
                ClientImportRow(
                    batch=batch,
                    row_key=idx,
                    full_name=full_name or "",
                    nif=nif or nif_raw or "",
                    phone=phone or "",
                    email=email or "",
                    address_line1=address or "",
                    postal_code=postal_code or "",
                    city=city or "",
                    county=county or "",
                    district=district or "",
                    valid_vat=is_valid_nif if validate_nif else bool(nif),
                    missing_email=missing_email,
                    duplicate_in_file=duplicate_in_file,
                    exists_in_db=exists_in_db,
                )
            )
        ClientImportRow.objects.bulk_create(bulk, batch_size=500)
        log_audit_event(
            category="client_import_batch",
            action="preview",
            request=request,
            instance=batch,
            source="client_import",
            message="Pré-visualização de importação criada.",
            after=_client_import_batch_snapshot(batch),
            metadata={
                "row_count": len(bulk),
                "validate_nif": validate_nif,
                "filename": upload.name,
            },
        )
        return batch

    if request.method == "POST":
        blocked, retry_after = check_rate_limit(
            request,
            name="backoffice_import_ip_minute",
            limit=20,
            window=60,
            by_ip=True,
        )
        if blocked:
            messages.error(request, "Demasiadas tentativas. Tenta novamente em alguns minutos.")
            response = redirect("client_import")
            response.status_code = 429
            response["Retry-After"] = str(retry_after)
            return response
        action = (request.POST.get("action") or "").strip()
        validate_nif = (request.POST.get("validate_nif") or "0") == "1"
        if action == "preview":
            upload = request.FILES.get("csv_file")
            if not upload:
                messages.error(request, "Seleciona um ficheiro CSV.")
                return redirect("client_import")
            try:
                batch = parse_csv_to_batch(upload, validate_nif)
            except ValueError as exc:
                messages.error(request, str(exc))
                return redirect("client_import")
            request.session["client_import_batch_id"] = batch.id
            request.session["import_selected_ids"] = []
            return redirect("client_import")

        batch_id = request.POST.get("batch_id") or batch_id
        if not batch_id:
            return redirect("client_import")
        batch = ClientImportBatch.objects.filter(id=batch_id).first()
        if not batch:
            return redirect("client_import")

        q = (request.POST.get("q") or "").strip()
        only_missing_email = request.POST.get("only_missing_email") == "1"
        only_valid_vat = request.POST.get("only_valid_vat") == "1"
        only_duplicates = request.POST.get("only_duplicates") == "1"
        page_size = request.POST.get("page_size") or "5"
        redirect_qs = urlencode(
            {
                "q": q,
                "page_size": page_size,
                "only_missing_email": "1" if only_missing_email else "0",
                "only_valid_vat": "1" if only_valid_vat else "0",
                "only_duplicates": "1" if only_duplicates else "0",
            }
        )

        rows_qs = ClientImportRow.objects.filter(batch=batch)
        if q:
            rows_qs = apply_terms_filter(
                rows_qs,
                q,
                ["full_name__icontains", "nif__icontains", "phone__icontains"],
            )
        if only_missing_email:
            rows_qs = rows_qs.filter(missing_email=True)
        if only_valid_vat:
            rows_qs = rows_qs.filter(valid_vat=True)
        if only_duplicates:
            rows_qs = rows_qs.filter(Q(duplicate_in_file=True) | Q(exists_in_db=True))

        if action == "clear":
            request.session["import_selected_ids"] = []
            return redirect(f"{reverse('client_import')}?{redirect_qs}")

        if action == "select_all_filtered":
            all_ids = list(rows_qs.values_list("id", flat=True))
            selected_ids = set(all_ids)
            request.session["import_selected_ids"] = list(selected_ids)
            return redirect(f"{reverse('client_import')}?{redirect_qs}")

        if action == "select_page" or action == "update_selection":
            row_ids_on_page = request.POST.get("row_ids_on_page", "")
            row_ids_on_page = [r for r in row_ids_on_page.split(",") if r]
            checked = request.POST.getlist("row_id")
            selected_ids = set(selected_ids)
            for rid in row_ids_on_page:
                selected_ids.discard(int(rid))
            for rid in checked:
                selected_ids.add(int(rid))
            request.session["import_selected_ids"] = list(selected_ids)
            return redirect(f"{reverse('client_import')}?{redirect_qs}")

        def import_rows(qs):
            created = 0
            updated = 0
            skipped = 0
            errors = 0
            processed_ids = []
            for row in qs:
                try:
                    processed_ids.append(row.id)
                    nif = "".join(ch for ch in (row.nif or "") if ch.isdigit())
                    if batch.validate_nif and nif and not row.valid_vat:
                        skipped += 1
                        continue
                    profile = ClientProfile.objects.filter(nif=nif).first() if nif else None
                    if not profile:
                        profile = find_existing_client_by_name_phone(row.full_name, row.phone)
                    if not profile and not nif:
                        skipped += 1
                        continue
                    if profile:
                        changed = False
                        if nif and not profile.nif:
                            profile.nif = nif
                            changed = True
                        if row.full_name and not profile.full_name:
                            profile.full_name = row.full_name
                            changed = True
                        if row.phone and not profile.phone:
                            profile.phone = row.phone
                            changed = True
                        if row.address_line1 and not profile.address_line1:
                            profile.address_line1 = row.address_line1
                            changed = True
                        if row.postal_code and not profile.postal_code:
                            profile.postal_code = row.postal_code
                            changed = True
                        if row.city and not profile.city:
                            profile.city = row.city
                            changed = True
                        if row.county and not profile.county:
                            profile.county = row.county
                            changed = True
                        if row.district and not profile.district:
                            profile.district = row.district
                            changed = True
                        if row.email and profile.user and not profile.user.email:
                            profile.user.email = row.email
                            profile.user.save(update_fields=["email"])
                        if changed:
                            profile.save()
                            updated += 1
                        else:
                            skipped += 1
                    else:
                        ClientProfile.objects.create(
                            user=None,
                            full_name=row.full_name or "—",
                            nif=nif,
                            phone=row.phone or "",
                            address_line1=row.address_line1 or "",
                            postal_code=row.postal_code or "",
                            city=row.city or "",
                            county=row.county or "",
                            district=row.district or "",
                        )
                        created += 1
                except Exception:
                    errors += 1
            log = ClientImportLog.objects.create(
                created_by=request.user,
                file_name=batch.original_filename,
                created_count=created,
                updated_count=updated,
                skipped_count=skipped,
                error_count=errors,
                summary=f"Criados {created}, atualizados {updated}, ignorados {skipped}, erros {errors}",
            )
            return log, created, updated, skipped, errors, processed_ids

        if action == "import_selected":
            row_ids_on_page = request.POST.get("row_ids_on_page", "")
            row_ids_on_page = [r for r in row_ids_on_page.split(",") if r]
            checked = request.POST.getlist("row_id")
            selected_ids = set(selected_ids)
            for rid in row_ids_on_page:
                selected_ids.discard(int(rid))
            for rid in checked:
                selected_ids.add(int(rid))
            request.session["import_selected_ids"] = list(selected_ids)
            ids = list(selected_ids)
            if not ids:
                messages.error(request, "Seleciona pelo menos uma linha para importar.")
                return redirect("client_import")
            rows = ClientImportRow.objects.filter(batch=batch, id__in=ids)
            log, created, updated, skipped, errors, processed_ids = import_rows(rows)
            log_audit_event(
                category="client_import",
                action="import_selected",
                request=request,
                instance=log,
                source="client_import",
                message="Importação de clientes a partir da seleção.",
                after=_client_import_log_snapshot(log),
                metadata={
                    "batch": _client_import_batch_snapshot(batch),
                    "selected_ids_count": len(ids),
                    "processed_ids_count": len(processed_ids),
                },
            )
            messages.success(
                request,
                f"Importação concluída: {created} criados, {updated} atualizados, {skipped} ignorados, {errors} erros.",
            )
            return redirect("client_import")

        if action == "import_filtered":
            if not rows_qs.exists():
                messages.error(request, "Não há resultados filtrados para importar.")
                return redirect("client_import")
            log, created, updated, skipped, errors, processed_ids = import_rows(rows_qs)
            log_audit_event(
                category="client_import",
                action="import_filtered",
                request=request,
                instance=log,
                source="client_import",
                message="Importação de clientes a partir do filtro atual.",
                after=_client_import_log_snapshot(log),
                metadata={
                    "batch": _client_import_batch_snapshot(batch),
                    "processed_ids_count": len(processed_ids),
                    "filters": {
                        "q": q,
                        "only_missing_email": only_missing_email,
                        "only_valid_vat": only_valid_vat,
                        "only_duplicates": only_duplicates,
                    },
                },
            )
            messages.success(
                request,
                f"Importação concluída: {created} criados, {updated} atualizados, {skipped} ignorados, {errors} erros.",
            )
            return redirect("client_import")

    logs = ClientImportLog.objects.all()[:10]
    batch = ClientImportBatch.objects.filter(id=batch_id).first() if batch_id else None
    q = (request.GET.get("q") or "").strip()
    only_missing_email = request.GET.get("only_missing_email") == "1"
    only_valid_vat = request.GET.get("only_valid_vat") == "1"
    only_duplicates = request.GET.get("only_duplicates") == "1"
    page_size = request.GET.get("page_size") or "5"
    try:
        page_size = int(page_size)
    except Exception:
        page_size = 5
    if page_size not in (5, 10, 15, 25):
        page_size = 5

    rows_page = None
    total_filtered = 0
    if batch:
        rows_qs = ClientImportRow.objects.filter(batch=batch)
        if q:
            rows_qs = apply_terms_filter(
                rows_qs,
                q,
                ["full_name__icontains", "nif__icontains", "phone__icontains"],
            )
        if only_missing_email:
            rows_qs = rows_qs.filter(missing_email=True)
        if only_valid_vat:
            rows_qs = rows_qs.filter(valid_vat=True)
        if only_duplicates:
            rows_qs = rows_qs.filter(Q(duplicate_in_file=True) | Q(exists_in_db=True))
        total_filtered = rows_qs.count()
        rows_page = Paginator(rows_qs, page_size).get_page(request.GET.get("page") or 1)

    return render(
        request,
        "core/client_import.html",
        {
            "logs": logs,
            "batch": batch,
            "rows_page": rows_page,
            "q": q,
            "only_missing_email": only_missing_email,
            "only_valid_vat": only_valid_vat,
            "only_duplicates": only_duplicates,
            "page_size": page_size,
            "page_sizes": [5, 10, 15, 25],
            "selected_count": len(selected_ids),
            "selected_ids": selected_ids,
            "total_filtered": total_filtered,
        },
    )
